"""
Verify imported contacts — compare Unimarketing list contacts with original xlsx emails.

New approach: Read emails from xlsx, query each by email+listId via API.
This avoids the unstable pagination entirely.

Public API (for GUI integration):
  get_list_info(list_id)              -> dict | None
  verify_list_import(list_id, xlsx, dir, logger, is_formal) -> tuple[bool, str]
  find_lists_by_sn(sn)                -> list[tuple[list_id, title, type]]
"""
import os
import sys
import time
from xml.etree import ElementTree as ET

import requests
from requests.auth import HTTPBasicAuth

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Unimarketing API constants
# ---------------------------------------------------------------------------
HOST = "http://services.unimarketing.com.cn"
API_KEY = "customersupport"
API_SECRET = "/CUkafFTgALhtSSZn9KcZ1hw4lI="

ATOM_NS = "http://www.w3.org/2005/Atom"
UM_NS = "http://www.unimarketing.com.cn/xmlns/"
OS_NS = "http://a9.com/-/spec/opensearchrss/1.0/"


def _api_headers():
    return {
        "Accept": "application/atom+xml",
        "Authorization": "OAuth",
    }


def _api_auth():
    return HTTPBasicAuth(API_KEY, API_SECRET)


def _atom(tag: str) -> str:
    return "{" + ATOM_NS + "}" + tag


def _os(tag: str) -> str:
    return "{" + OS_NS + "}" + tag


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------
def get_list_info(list_id: str) -> dict | None:
    """GET /list/{id}/ — return dict with list fields."""
    params = {"apikey": API_KEY, "method": "get", "alt": "atom"}
    resp = requests.get(
        f"{HOST}/list/{list_id}/",
        params=params,
        headers=_api_headers(),
        auth=_api_auth(),
        timeout=30,
    )
    if resp.status_code != 200:
        return None

    root = ET.fromstring(resp.content)
    info = {}

    def get_text(tag):
        el = root.find(_atom(tag))
        return el.text.strip() if el is not None and el.text else None

    def get_um_text(tag):
        el = root.find("{" + UM_NS + "}" + tag)
        return el.text.strip() if el is not None and el.text else None

    info["title"] = get_text("title") or ""
    info["activeCount"] = int(get_um_text("activeCount")) if get_um_text("activeCount") else 0
    info["unsubscribeCount"] = int(get_um_text("unsubscribeCount")) if get_um_text("unsubscribeCount") else 0
    info["invalidateCount"] = int(get_um_text("invalidateCount")) if get_um_text("invalidateCount") else 0
    info["unconfirmCount"] = int(get_um_text("unconfirmCount")) if get_um_text("unconfirmCount") else 0
    return info


def _query_contact_by_email(email: str, list_id: str) -> bool:
    """Check if email exists in list. Returns True if found."""
    q = f"[email={email},listId={list_id}]"
    params = {
        "apikey": API_KEY,
        "method": "get",
        "alt": "atom",
        "q": q,
    }
    resp = requests.get(
        f"{HOST}/contact/",
        params=params,
        headers=_api_headers(),
        auth=_api_auth(),
        timeout=30,
    )
    if resp.status_code != 200:
        return False
    root = ET.fromstring(resp.content)
    entries = root.findall(_atom("entry"))
    return len(entries) > 0


def find_duplicates(items: list[str]) -> dict[str, int]:
    """Find duplicate strings (case-insensitive). Returns {lower_item: count} for items appearing >1."""
    seen: dict[str, int] = {}
    for item in items:
        k = item.lower()
        seen[k] = seen.get(k, 0) + 1
    return {k: v for k, v in seen.items() if v > 1}


def find_lists_by_sn(sn: str) -> list[tuple[str, str, str]]:
    """
    Search lists by SN number in title (fuzzy match — strips prefix, searches digits).
    Returns list of (list_id, title, type) where type is 'formal' or 'test'.
    """
    # Extract digits for fuzzy match — "SN-56287", "56287", "SN 56287" all become "56287"
    digits = "".join(c for c in sn if c.isdigit())
    q = digits if digits else sn
    params = {
        "apikey": API_KEY,
        "method": "get",
        "alt": "atom",
        "field": "title",
        "q": q,
    }
    resp = requests.get(
        f"{HOST}/list/",
        params=params,
        headers=_api_headers(),
        auth=_api_auth(),
        timeout=30,
    )
    if resp.status_code != 200:
        return []

    root = ET.fromstring(resp.content)
    results = []
    for entry in root.findall(_atom("entry")):
        tid = entry.find(_atom("id"))
        title = entry.find(_atom("title"))
        if tid is None or title is None:
            continue

        lid = tid.text.rsplit("/", 1)[-1]
        t = title.text or ""
        t_lower = t.lower()

        if t_lower.startswith("test_"):
            results.append((lid, t, "test"))
        else:
            # Anything not starting with "test_" is treated as formal
            results.append((lid, t, "formal"))

    return results


def read_xlsx_emails(xlsx_path: str) -> list[str]:
    """Read xlsx, extract all non-empty Email values."""
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    header = rows[0]
    email_idx = None
    for i, col in enumerate(header):
        if col is not None and str(col).strip().lower() == "email":
            email_idx = i
            break
    if email_idx is None:
        return []

    emails = []
    for row in rows[1:]:
        if email_idx < len(row) and row[email_idx]:
            val = str(row[email_idx]).strip()
            if val:
                emails.append(val)
    return emails


# ---------------------------------------------------------------------------
# Compare logic
# ---------------------------------------------------------------------------
def compare_emails(list_emails: list[str], xlsx_emails: list[str]) -> dict:
    """Compare two email sets, return report dict."""
    list_set = set(e.lower() for e in list_emails)
    xlsx_set = set(e.lower() for e in xlsx_emails)

    return {
        "list_total": len(list_emails),
        "list_unique": len(list_set),
        "xlsx_total": len(xlsx_emails),
        "xlsx_unique": len(xlsx_set),
        "match_count": len(xlsx_set & list_set),
        "missing_in_list": sorted(xlsx_set - list_set),
        "extra_in_list": sorted(list_set - xlsx_set),
        "pass": not (xlsx_set - list_set) and not (list_set - xlsx_set),
    }


# ---------------------------------------------------------------------------
# Full verify pipeline (for GUI import-verify)
# ---------------------------------------------------------------------------
def verify_list_import(
    list_id: str,
    xlsx_path: str,
    save_dir: str,
    logger,
    is_formal: bool = False,
) -> tuple[bool, str]:
    """
    Full verify: read emails from xlsx, query each by API in the list, compare.
    Returns (passed: bool, result_message: str).
    """
    if not hasattr(logger, "log") or not callable(getattr(logger, "log", None)):
        _orig_call = logger

        class _L:
            def log(self, msg):
                _orig_call(msg)

        logger = _L()

    from concurrent.futures import ThreadPoolExecutor, as_completed

    # 1. List info
    logger.log("[VERIFY] Fetching list info...")
    info = get_list_info(list_id)
    if not info:
        return False, "Could not get list info."
    list_title = info["title"]
    logger.log(f"[VERIFY] List: {list_title} (id={list_id})")

    # 2. Read xlsx emails
    logger.log("[VERIFY] Reading xlsx emails...")
    xlsx_emails = read_xlsx_emails(xlsx_path)
    logger.log(f"[VERIFY] XLSX emails: {len(xlsx_emails)}")

    # 3. Query each email from API (parallel, no pagination)
    logger.log("[VERIFY] Checking emails in list (parallel)...")
    found_emails = set()
    missing_emails = []
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=10) as pool:
        futs = {
            pool.submit(_query_contact_by_email, email, list_id): email
            for email in xlsx_emails
        }
        for fut in as_completed(futs):
            email = futs[fut]
            try:
                if fut.result():
                    found_emails.add(email.lower())
                else:
                    missing_emails.append(email)
            except Exception as e:
                logger.log(f"[VERIFY] ERROR checking {email}: {e}")

    elapsed = time.time() - start_time
    logger.log(f"[VERIFY] Checked {len(xlsx_emails)} emails in {elapsed:.1f}s")
    logger.log(f"[VERIFY] Found in list: {len(found_emails)}")
    logger.log(f"[VERIFY] Missing: {len(missing_emails)}")

    # 4. Save found emails
    os.makedirs(save_dir, exist_ok=True)
    safe_title = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in list_title)
    filename = f"{safe_title}_emails.txt"
    filepath = os.path.join(save_dir, filename)

    all_found = sorted(found_emails)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(f"# List ID: {list_id}\n")
        f.write(f"# Title: {list_title}\n")
        f.write(f"# Total: {len(all_found)}\n")
        f.write(f"# Generated by verify_list_contacts.py\n\n")
        for e in all_found:
            f.write(e + "\n")
    logger.log(f"[VERIFY] Saved to: {filepath}")

    # 5. Compare
    xlsx_lower = set(e.lower() for e in xlsx_emails)
    match_count = len(xlsx_lower & found_emails)

    total_in_list = (
        info["activeCount"] + info["unsubscribeCount"]
        + info["invalidateCount"] + info["unconfirmCount"]
    )

    report_lines = [
        "",
        f"[VERIFY] ={'='*58}",
        f"[VERIFY]   {list_title}",
        f"[VERIFY] ={'='*58}",
        f"[VERIFY]   List total     : {total_in_list}",
        f"[VERIFY]   XLSX emails    : {len(xlsx_emails)}",
        f"[VERIFY]   In list        : {len(found_emails)}",
        f"[VERIFY]   Matched        : {match_count}",
    ]

    for line in report_lines:
        logger.log(line)

    if missing_emails:
        logger.log(f"[VERIFY]   MISSING ({len(missing_emails)}):")
        for e in sorted(missing_emails):
            logger.log(f"[VERIFY]     - {e}")

    extra_in_list = found_emails - xlsx_lower
    if extra_in_list:
        logger.log(f"[VERIFY]   EXTRA ({len(extra_in_list)}):")
        for e in sorted(extra_in_list):
            logger.log(f"[VERIFY]     + {e}")

    passed = not missing_emails
    if passed:
        logger.log(f"[VERIFY]   RESULT: PASS")
        logger.log(f"[VERIFY] ={'='*58}")
        msg = (
            f"验证成功！\n"
            f"List: {list_title}\n"
            f"XLSX: {len(xlsx_emails)} | 全部在列表中"
        )
        return True, msg
    else:
        logger.log(f"[VERIFY]   RESULT: FAIL")
        logger.log(f"[VERIFY] ={'='*58}")
        msg = (
            f"验证失败！\n"
            f"List: {list_title}\n"
            f"XLSX: {len(xlsx_emails)} | 列表中找到: {len(found_emails)}\n"
            f"匹配: {match_count} | 缺失: {len(missing_emails)}\n\n"
            f"请手动检查 Unimarketing 列表。"
        )
        return False, msg


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Verify Unimarketing list contacts against xlsx emails"
    )
    parser.add_argument("--list-id", default=None, help="Unimarketing list ID")
    parser.add_argument("--sn", default=None, help="SN number (auto-discovers lists)")
    parser.add_argument("--xlsx", default=None, help="Path to original xlsx file")
    parser.add_argument("--list-type", default="formal", choices=["formal", "test"])
    parser.add_argument("--probe", action="store_true", help="Probe list info")
    args = parser.parse_args()

    list_id = args.list_id
    if args.sn and not list_id:
        found = find_lists_by_sn(args.sn)
        matching = [(lid, t, ty) for lid, t, ty in found if ty == args.list_type]
        if matching:
            list_id = max(matching, key=lambda x: int(x[0]))[0]
            print(f"Selected: listId={list_id} title={matching[0][1]}")
        else:
            print(f"No {args.list_type} list found for SN '{args.sn}'", file=sys.stderr)
            sys.exit(1)

    if not list_id:
        parser.error("--list-id or --sn required")

    if args.probe:
        info = get_list_info(list_id)
        if info:
            for k, v in info.items():
                print(f"  {k}: {v}")
        else:
            print("Cannot get list info.")
        return

    if not args.xlsx:
        parser.error("--xlsx required (or use --probe)")

    # Logger — simple print logger
    class PrintLogger:
        def log(self, msg): print(msg)

    ok, msg = verify_list_import(
        list_id, args.xlsx,
        os.path.join(os.path.dirname(args.xlsx), "listverify"),
        PrintLogger(),
    )
    if not ok:
        print(msg, file=sys.stderr)
        sys.exit(1)
    print(msg)
    sys.exit(0)


if __name__ == "__main__":
    main()
