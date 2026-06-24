"""
Deep list verification — compare Unimarketing list contacts with xlsx, field by field.

New approach: Read emails from xlsx, query each contact by email+listId via API.
This avoids the unstable pagination entirely.

Steps:
  1. Read xlsx emails
  2. Parallel query each email: GET /contact/?q=[email=X,listId=Y]
  3. Collect all found contacts, save to CSV
  4. Compare fields xlsx vs API results

Public API (for GUI integration):
  deep_verify(list_id, xlsx_path, save_dir, logger, max_workers) -> tuple[bool, str]
"""
import csv
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from xml.etree import ElementTree as ET

import requests
from requests.auth import HTTPBasicAuth

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Unimarketing API constants
# ---------------------------------------------------------------------------
HOST = "http://services.unimarketing.com.cn"
API_KEY = "customersupport"
API_SECRET = "/CUkafFTgALhtSSZn9KcZ1hw4lI="

ATOM_NS = "http://www.w3.org/2005/Atom"
UM_NS = "http://www.unimarketing.com.cn/xmlns/"


def _headers():
    return {
        "Accept": "application/atom+xml",
        "Authorization": "OAuth",
    }


def _auth():
    return HTTPBasicAuth(API_KEY, API_SECRET)


def _atom(tag):
    return "{" + ATOM_NS + "}" + tag


def _um(tag):
    return "{" + UM_NS + "}" + tag


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------
def get_list_info(list_id: str) -> dict | None:
    """GET /list/{id}/ — return dict with list fields."""
    params = {"apikey": API_KEY, "method": "get", "alt": "atom"}
    resp = requests.get(
        f"{HOST}/list/{list_id}/",
        params=params,
        headers=_headers(),
        auth=_auth(),
        timeout=30,
    )
    if resp.status_code != 200:
        return None

    root = ET.fromstring(resp.content)

    def _t(tag):
        el = root.find(_atom(tag))
        return el.text.strip() if el is not None and el.text else None

    def _um_t(tag):
        el = root.find(_um(tag))
        return el.text.strip() if el is not None and el.text else None

    return {
        "title": _t("title") or "",
        "activeCount": int(_um_t("activeCount")) if _um_t("activeCount") else 0,
        "unsubscribeCount": int(_um_t("unsubscribeCount")) if _um_t("unsubscribeCount") else 0,
        "invalidateCount": int(_um_t("invalidateCount")) if _um_t("invalidateCount") else 0,
        "unconfirmCount": int(_um_t("unconfirmCount")) if _um_t("unconfirmCount") else 0,
    }


def query_contact_by_email(email: str, list_id: str) -> dict | None:
    """Query a single contact by email in a list.
    Returns {email, attributes} or None if not found.
    """
    q = f"[email={email},listId={list_id}]"
    params = {
        "apikey": API_KEY,
        "method": "get",
        "alt": "atom",
        "q": q,
    }
    resp = requests.get(
        f"{HOST}/contact/",
        params=params,
        headers=_headers(),
        auth=_auth(),
        timeout=30,
    )
    if resp.status_code != 200:
        return None

    root = ET.fromstring(resp.content)
    entries = root.findall(_atom("entry"))
    if not entries:
        return None

    entry = entries[0]
    c = {}

    email_el = entry.find(_atom("email"))
    if email_el is not None and email_el.text:
        c["email"] = email_el.text.strip()

    attrs = {}
    for attr in entry.findall(_um("attribute")):
        name = attr.get("name", "")
        label = attr.get("label", name)
        if name:
            csv_col = label if name.startswith(("Token", "SubId")) or name in (
                "SubIdT", "SubIdH", "SubIdF", "SubIdI", "SubIdS", "SubIdE", "SubIdG",
                "SubIdN", "SubIdTEN", "SubIdL", "SubIdW", "SubIdR", "SubIdO",
                "TokenT", "TokenH", "TokenF", "TokenI", "TokenS", "TokenE", "TokenG",
                "TokenN", "TokenTEN", "TokenL", "TokenW", "TokenR", "TokenO", "TokenV") else name
            attrs[csv_col] = (attr.text or "").strip()

    if attrs:
        c["attributes"] = attrs
    if c.get("email"):
        return c
    return None


def save_contacts_csv(contacts: list[dict], list_id: str, sn: str, save_dir: str) -> str:
    """Save contacts to UTF-8 CSV file. Returns file path."""
    os.makedirs(save_dir, exist_ok=True)
    now = time.strftime("%Y%m%d_%H%M%S")
    filename = f"ContactInfo_exported_{sn}_{now}.csv"
    filepath = os.path.join(save_dir, filename)

    if not contacts:
        with open(filepath, "w", encoding="utf-8", newline="") as f:
            f.write("")
        return filepath

    all_labels = []
    seen = set()
    for c in contacts:
        for label in (c.get("attributes") or {}).keys():
            if label not in seen:
                all_labels.append(label)
                seen.add(label)

    with open(filepath, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["email"] + all_labels)
        for c in contacts:
            attrs = c.get("attributes") or {}
            row = [c.get("email", "")] + [attrs.get(l, "") for l in all_labels]
            writer.writerow(row)

    return filepath


def load_csv_as_map(csv_path: str) -> tuple[list[str], dict[str, dict[str, str]]]:
    """Read CSV, return (headers, Map<email, {header: value}>)."""
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        data = {}
        for row in reader:
            email = row.get("email", "").strip().lower()
            if email:
                data[email] = {k: (row.get(k) or "").strip() for k in headers}
    return headers, data


def read_xlsx_as_map(xlsx_path: str) -> tuple[list[str], dict[str, dict[str, str]]]:
    """Read xlsx via openpyxl, all cells as String.
    Returns (headers, Map<email, {header: value}>).
    """
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], {}

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    email_idx = None
    for i, h in enumerate(headers):
        if h.lower() == "email":
            email_idx = i
            break

    if email_idx is None:
        return headers, {}

    data = {}
    for row in rows[1:]:
        values = [str(c).strip() if c is not None else "" for c in row]
        email = values[email_idx].lower()
        if email:
            data[email] = {headers[i]: values[i] for i in range(len(headers))}

    return headers, data


def deep_compare(
    xlsx_headers: list[str],
    xlsx_data: dict[str, dict[str, str]],
    csv_headers: list[str],
    csv_data: dict[str, dict[str, str]],
) -> dict:
    """Deep compare xlsx vs API contacts."""
    xlsx_emails = set(xlsx_data.keys())
    csv_emails = set(csv_data.keys())

    match_emails = xlsx_emails & csv_emails
    missing_in_list = sorted(xlsx_emails - csv_emails)
    extra_in_list = sorted(csv_emails - xlsx_emails)

    csv_headers_lower = {h.lower().strip(): h for h in csv_headers}

    field_mismatches = {}
    total_fields_compared = 0
    total_fields_matched = 0
    detail_differences = []

    for email in sorted(match_emails):
        x_row = xlsx_data[email]
        c_row = csv_data[email]

        for xh in xlsx_headers:
            xh_lower = xh.lower().strip()
            if xh_lower == "email":
                continue

            csv_match = csv_headers_lower.get(xh_lower)
            c_val = c_row.get(csv_match, "") if csv_match else ""
            x_val = x_row.get(xh, "")

            total_fields_compared += 1

            x_val_stripped = x_val.strip()
            c_val_stripped = c_val.strip()

            matched = False
            if x_val_stripped == c_val_stripped:
                matched = True
            elif c_val_stripped:
                if ", " in x_val_stripped:
                    display_part = x_val_stripped.split(", ", 1)[1].strip()
                    if display_part == c_val_stripped:
                        matched = True
                if ", " in c_val_stripped:
                    display_part = c_val_stripped.split(", ", 1)[1].strip()
                    if display_part == x_val_stripped:
                        matched = True

            if matched:
                total_fields_matched += 1
            else:
                field_mismatches[xh] = field_mismatches.get(xh, 0) + 1
                detail_differences.append({
                    "email": email,
                    "field": xh,
                    "xlsx": x_val_stripped,
                    "csv": c_val_stripped,
                })

    return {
        "xlsx_count": len(xlsx_data),
        "csv_count": len(csv_data),
        "match_count": len(match_emails),
        "missing_in_list": missing_in_list,
        "extra_in_list": extra_in_list,
        "email_count_pass": len(missing_in_list) == 0 and len(extra_in_list) == 0,
        "fields_compared": total_fields_compared,
        "fields_matched": total_fields_matched,
        "field_mismatches": field_mismatches,
        "detail_differences": detail_differences,
        "xlsx_headers": xlsx_headers,
        "csv_headers": csv_headers,
    }


def deep_verify(
    list_id: str,
    xlsx_path: str,
    save_dir: str,
    logger,
    max_workers: int = 10,
) -> tuple[bool, str]:
    """Full deep verify: query each xlsx email by API, compare field by field."""
    if not hasattr(logger, "log") or not callable(getattr(logger, "log", None)):
        _orig = logger

        class _L:
            def log(self, msg):
                _orig(msg)

        logger = _L()

    sn = ""
    sn_match = re.search(r"SN[- ]?(\d+)", xlsx_path, re.IGNORECASE)
    if sn_match:
        sn = f"SN-{sn_match.group(1)}"

    # Step 1: Read xlsx
    logger.log("[DEEP] Reading xlsx...")
    xlsx_headers, xlsx_data = read_xlsx_as_map(xlsx_path)
    logger.log(f"[DEEP] XLSX: {len(xlsx_data)} emails")

    # Step 2: Query each email from API (parallel, no pagination issues)
    logger.log(f"[DEEP] Querying {len(xlsx_data)} emails via API (workers={max_workers})...")
    xlsx_emails = list(xlsx_data.keys())
    contacts = []
    found_count = 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futs = {
            pool.submit(query_contact_by_email, email, list_id): email
            for email in xlsx_emails
        }
        for fut in as_completed(futs):
            email = futs[fut]
            try:
                contact = fut.result()
                if contact:
                    contacts.append(contact)
                    found_count += 1
            except Exception as e:
                logger.log(f"[DEEP] ERROR querying {email}: {e}")

    elapsed = time.time() - start_time
    logger.log(f"[DEEP] Found {found_count}/{len(xlsx_emails)} in list ({elapsed:.1f}s)")

    # Step 3: Save to CSV
    logger.log("[DEEP] Saving contacts to CSV...")
    csv_path = save_contacts_csv(contacts, list_id, sn, save_dir)
    logger.log(f"[DEEP] Saved CSV: {os.path.basename(csv_path)}")

    # Step 4: Load CSV
    csv_headers, csv_data = load_csv_as_map(csv_path)

    # Step 5: Deep compare
    logger.log("[DEEP] Comparing field by field...")
    result = deep_compare(xlsx_headers, xlsx_data, csv_headers, csv_data)

    # Build report
    lines = []
    lines.append("")
    lines.append("=" * 60)
    lines.append(f"  DEEP VERIFY: {sn}")
    lines.append("=" * 60)
    lines.append(f"  XLSX contacts   : {result['xlsx_count']}")
    lines.append(f"  In list         : {result['csv_count']}")
    lines.append(f"  Email matched   : {result['match_count']}")

    email_pass = result["email_count_pass"]
    lines.append(f"  Email check     : {'PASS' if email_pass else 'FAIL'}")

    if result["missing_in_list"]:
        lines.append(f"  Missing in list ({len(result['missing_in_list'])}):")
        for e in result["missing_in_list"]:
            lines.append(f"    - {e}")

    if result["extra_in_list"]:
        lines.append(f"  Extra in list ({len(result['extra_in_list'])}):")
        for e in result["extra_in_list"]:
            lines.append(f"    + {e}")

    lines.append("")
    lines.append("-" * 60)
    lines.append(f"  Fields compared : {result['fields_compared']}")
    lines.append(f"  Fields matched  : {result['fields_matched']}")

    field_pass = len(result["field_mismatches"]) == 0
    lines.append(f"  Field match     : {'PASS' if field_pass else 'FAIL'}")

    if result["field_mismatches"]:
        lines.append("")
        lines.append("  Field mismatches:")
        for field, count in sorted(result["field_mismatches"].items()):
            lines.append(f"    {field}: {count} mismatch(es)")

        lines.append("")
        lines.append("  Detail (first 20):")
        for d in result["detail_differences"][:20]:
            lines.append(f"    [{d['email']}] {d['field']}:")
            lines.append(f"      XLSX: {d['xlsx'][:80]}")
            lines.append(f"      API : {d['csv'][:80]}")

    overall = email_pass and field_pass
    lines.append("")
    lines.append("=" * 60)
    lines.append(f"  RESULT: {'PASS' if overall else 'FAIL'}")
    lines.append("=" * 60)

    for line in lines:
        logger.log(line)

    if overall:
        msg = (
            f"深验证成功！\n"
            f"SN: {sn}\n"
            f"联系人: {result['xlsx_count']} | 邮箱匹配: {result['match_count']}\n"
            f"字段比对: {result['fields_matched']}/{result['fields_compared']}\n"
            f"全部一致"
        )
    else:
        msg = (
            f"深验证完成！\n"
            f"SN: {sn}\n"
            f"联系人: {result['xlsx_count']} | 列表找到: {result['csv_count']}\n"
            f"邮箱匹配: {result['match_count']}/{result['xlsx_count']}\n"
            f"  缺失: {len(result['missing_in_list'])} | 多余: {len(result['extra_in_list'])}\n"
            f"字段: {result['fields_matched']}/{result['fields_compared']}\n"
        )
        if result["field_mismatches"]:
            mismatch_summary = ", ".join(
                f"{k}({v})" for k, v in list(result["field_mismatches"].items())[:5]
            )
            msg += f"字段差异: {mismatch_summary}\n"
        msg += "\n请查看日志获取详细信息。"

    return overall, msg


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------
def main():
    import argparse

    parser = argparse.ArgumentParser(description="Deep verify Unimarketing list vs xlsx")
    parser.add_argument("--list-id", required=True, help="Unimarketing list ID")
    parser.add_argument("--xlsx", required=True, help="Path to xlsx file")
    parser.add_argument("--workers", type=int, default=10, help="Parallel workers")
    parser.add_argument("--save-dir", default=None, help="Output directory for CSV")
    args = parser.parse_args()

    save_dir = args.save_dir or os.path.join(os.path.dirname(args.xlsx), "listverify")

    class PrintLogger:
        def log(self, msg): print(msg)

    logger = PrintLogger()
    passed, msg = deep_verify(args.list_id, args.xlsx, save_dir, logger, args.workers)
    print(msg)
    sys.exit(0 if passed else 1)


if __name__ == "__main__":
    main()
