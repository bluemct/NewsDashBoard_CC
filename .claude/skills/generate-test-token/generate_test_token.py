"""
Generate test token xlsx — picks two distinct rows with the most tokens filled,
replaces the Email column with test emails, saves as test_<original>.xlsx.

Usage:
    python generate_test_token.py <xlsx_path> [email1] [email2]

Arguments:
    xlsx_path  Token xlsx file path (required)
    email1     First test email (default: microsoft.163163@163.com)
    email2     Second test email (default: ma.chuntao@oe.21vianet.com)

Output:
    test_<original_name>.xlsx in the same directory as the input file.
"""
import os
import sys


def find_two_distinct_rows(ws):
    """Find two rows with most non-empty Token cells that have different Token values."""
    # Group by full token tuple, keep first occurrence of each distinct group
    seen = {}
    for r in range(2, ws.max_row + 1):
        tokens = tuple(ws.cell(row=r, column=c).value for c in range(2, 17))
        if tokens not in seen:
            seen[tokens] = r

    # Sort by number of non-empty tokens descending
    groups = []
    for tokens, rn in seen.items():
        filled = sum(1 for t in tokens if t is not None and str(t).strip())
        if filled > 0:
            groups.append((rn, filled))
    groups.sort(key=lambda x: x[1], reverse=True)

    if len(groups) < 2:
        print("Warning: less than 2 distinct rows found", file=sys.stderr)
        return [rn for rn, _ in groups[:2]]

    return [rn for rn, _ in groups[:2]]


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_test_token.py <xlsx_path> [email1] [email2]", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    email1 = sys.argv[2] if len(sys.argv) > 2 else "microsoft.163163@163.com"
    email2 = sys.argv[3] if len(sys.argv) > 3 else "ma.chuntao@oe.21vianet.com"

    if not os.path.isfile(xlsx_path):
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Find two distinct rows with most tokens filled
    rows = find_two_distinct_rows(ws)

    # Build output file path
    dir_name = os.path.dirname(xlsx_path)
    base_name = os.path.basename(xlsx_path)
    out_path = os.path.join(dir_name, "test_" + base_name)

    # Create new workbook
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Copy header row
    for c, cell in enumerate(ws[1], 1):
        new_ws.cell(row=1, column=c, value=cell.value)

    # Copy the two selected rows with replaced Email
    test_emails = [email1, email2]
    for i, orig_row in enumerate(rows):
        for c, cell in enumerate(ws[orig_row], 1):
            if c == 1:
                new_ws.cell(row=i + 2, column=c, value=test_emails[i])
            else:
                new_ws.cell(row=i + 2, column=c, value=cell.value)

    new_wb.save(out_path)
    wb.close()

    # Print summary
    filled = []
    for i in range(2, 4):
        count = sum(1 for c in range(2, 17) if new_ws.cell(row=i, column=c).value)
        filled.append(count)

    print(f"Source: {base_name}")
    print(f"Output: test_{base_name}")
    print(f"  Row 2: {test_emails[0]} (from original row {rows[0]}, {filled[0]}/15 tokens)")
    print(f"  Row 3: {test_emails[1]} (from original row {rows[1]}, {filled[1]}/15 tokens)")


if __name__ == "__main__":
    main()