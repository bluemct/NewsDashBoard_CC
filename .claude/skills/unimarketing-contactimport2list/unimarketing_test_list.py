"""
Unimarketing Contact List Import — Generate CSV from xlsx and import to a new list.

Two modes:
  --test   (default)  Pick rows with most tokens, replace Email with test emails
  --formal  Export all rows with original emails

Flow:
  1. Read xlsx → generate CSV (test or formal)
  2. Create Unimarketing list with Token/SubId attribute definitions
  3. Create import task (building)
  4. Submit contacts from CSV
  5. Execute import task
  6. Poll until import complete

Usage:
  # Test import (default)
  python unimarketing_test_list.py --xlsx "path/to/contacts.xlsx"

  # Formal import (all rows)
  python unimarketing_test_list.py --xlsx "path/to/contacts.xlsx" --formal

  # With explicit SN
  python unimarketing_test_list.py --xlsx "path/to/contacts.xlsx" --sn "SN-56230" --formal

Config files (project root):
  config.json — {"test_emails": ["a@b.com", "c@d.com"]}
"""
import argparse
import csv
import json
import os
import re
import sys
import time
import xml.sax.saxutils as sax
from datetime import datetime

import requests
from requests.auth import HTTPBasicAuth

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Unimarketing API constants
# ---------------------------------------------------------------------------
HOST = "http://services.unimarketing.com.cn"
API_KEY = "customersupport"
API_SECRET = "/CUkafFTgALhtSSZn9KcZ1hw4lI="

ATOM_NS = "http://www.w3.org/2005/Atom"
UM_NS = "http://www.unimarketing.com.cn/xmlns/"

# Token/SubId header -> system attribute name mapping
TOKEN_MAP = {
    "Token1": "Token", "Token2": "TokenT", "Token3": "TokenH", "Token4": "TokenF",
    "Token5": "TokenI", "Token6": "TokenS", "Token7": "TokenE", "Token8": "TokenG",
    "Token9": "TokenN", "Token10": "TokenTEN", "Token11": "TokenL", "Token12": "TokenW",
    "Token13": "TokenR", "Token14": "TokenO", "Token15": "TokenV",
}

SUBID_MAP = {
    "SubId1": "SubId", "SubId2": "SubIdT", "SubId3": "SubIdH", "SubId4": "SubIdF",
    "SubId5": "SubIdI", "SubId6": "SubIdS", "SubId7": "SubIdE", "SubId8": "SubIdG",
    "SubId9": "SubIdN", "SubId10": "SubIdTEN", "SubId11": "SubIdL", "SubId12": "SubIdW",
    "SubId13": "SubIdR", "SubId14": "SubIdO", "SubId15": "SubIdV",
}

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

if getattr(sys, "frozen", False):
    _SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    _SCRIPT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

_CONFIG_PATH = os.path.join(_SCRIPT_DIR, "config.json")


def _load_config() -> dict:
    defaults = {"test_emails": ["ma.chuntao@oe.21vianet.com", "microsoft.163163@163.com"]}
    if os.path.isfile(_CONFIG_PATH):
        with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
            defaults.update(json.load(f))
    return defaults

# ---------------------------------------------------------------------------
# Step A — xlsx → CSV
# ---------------------------------------------------------------------------


def _read_xlsx_rows(xlsx_path: str) -> tuple[list[str], list[list[str]]]:
    """Read first sheet of xlsx, return (header, data_rows)."""
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    wb.close()
    if len(rows) < 2:
        print("[CSV] xlsx has no data rows", file=sys.stderr)
        sys.exit(1)
    return rows[0], rows[1:]


def _find_attr_cols(header: list[str]) -> list[int]:
    """Return indices of Token/SubId columns."""
    return [i for i, col in enumerate(header) if col.strip().lower().startswith(("token", "subid"))]


def generate_test_csv(xlsx_path: str, output_dir: str, test_emails: list[str]) -> str:
    """
    Read xlsx, pick rows with most tokens filled, replace Email with test emails,
    write test_*.csv (GBK encoding). Returns CSV path.
    """
    header, data = _read_xlsx_rows(xlsx_path)

    email_idx = None
    for i, col in enumerate(header):
        if col.strip().lower() == "email":
            email_idx = i
            break

    attr_cols = _find_attr_cols(header)
    scored = []
    for i, row in enumerate(data):
        score = sum(1 for idx in attr_cols if idx < len(row) and row[idx].strip())
        if score > 0:
            scored.append((score, i, row))
    scored.sort(reverse=True)

    test_count = max(len(test_emails), 2)
    selected = []
    idx = 0
    while len(selected) < test_count:
        if idx < len(scored):
            selected.append(list(scored[idx][2]))
        elif len(data) > len(selected):
            selected.append(list(data[len(selected)]))
        else:
            selected.append(["" for _ in header])
        idx += 1

    if email_idx is not None:
        for i, row in enumerate(selected):
            if i < len(test_emails):
                row[email_idx] = test_emails[i]

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    csv_path = os.path.join(output_dir, f"test_{base}.csv")
    with open(csv_path, "w", encoding="gbk", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(selected)

    print(f"[CSV] test CSV: {os.path.basename(csv_path)} ({len(selected)} rows)")
    return csv_path


def generate_formal_csv(xlsx_path: str, output_dir: str) -> str:
    """
    Read xlsx, export all rows (original emails preserved), write formal_*.csv (GBK).
    Returns CSV path.
    """
    header, data = _read_xlsx_rows(xlsx_path)

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    csv_path = os.path.join(output_dir, f"formal_{base}.csv")
    with open(csv_path, "w", encoding="gbk", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data)

    print(f"[CSV] formal CSV: {os.path.basename(csv_path)} ({len(data)} rows)")
    return csv_path

# ---------------------------------------------------------------------------
# Step B — Parse CSV headers to build attribute definitions
# ---------------------------------------------------------------------------


def get_attr_mapping(header: list[str]) -> list[tuple[str, str]]:
    """
    Return list of (header_label, system_attr_name) for Token/SubId columns.
    Case-insensitive match — handles token1, Token1, TOKEN1, etc.
    """
    mapping = {}
    for n in range(1, 16):
        mapping[f"token{n}"] = TOKEN_MAP[f"Token{n}"]
        mapping[f"subid{n}"] = SUBID_MAP[f"SubId{n}"]

    result = []
    for col in header:
        stripped = col.strip()
        lower = stripped.lower()
        if lower in mapping:
            result.append((stripped, mapping[lower]))
    return result

# ---------------------------------------------------------------------------
# Step C — Unimarketing API helpers
# ---------------------------------------------------------------------------


def _api_headers() -> dict:
    return {
        "Content-Type": "application/atom+xml; charset=utf-8",
        "Accept": "application/atom+xml",
        "Authorization": "OAuth",
    }


def _api_auth():
    return HTTPBasicAuth(API_KEY, API_SECRET)


def parse_xml_response(resp) -> str:
    """Decode response — tries GBK first (server encoding), then UTF-8."""
    try:
        return resp.content.decode("gbk")
    except (UnicodeDecodeError, LookupError):
        return resp.content.decode("utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Step 1 — Create List
# ---------------------------------------------------------------------------


def create_list(list_title: str, attrs: list[tuple[str, str]]) -> str | None:
    """
    POST /list/ to create a contact list with attribute definitions.
    Returns listId on success, None on failure.
    """
    attr_xml = ""
    for sn, (label, sys_name) in enumerate(attrs, start=1):
        attr_xml += (
            f'  <um:attribute name="{sys_name}" label="{sax.escape(label)}" '
            f'visible="true" public="true" sn="{sn}" type="text"></um:attribute>\n'
        )

    body = (
        f'<entry xmlns="{ATOM_NS}" xmlns:um="{UM_NS}">\n'
        f'  <title>{sax.escape(list_title)}</title>\n'
        f'  <subtitle>{sax.escape(list_title)}</subtitle>\n'
        f'{attr_xml}'
        f'</entry>'
    )

    resp = requests.post(
        f"{HOST}/list/",
        params={"apikey": API_KEY, "method": "post", "alt": "atom"},
        headers=_api_headers(),
        auth=_api_auth(),
        data=body.encode("utf-8"),
        timeout=30,
    )

    if resp.status_code not in (200, 201):
        text = parse_xml_response(resp)
        print(f"[API] create list failed: {resp.status_code} — {text}", file=sys.stderr)
        return None

    root = __import__("xml.etree.ElementTree", fromlist=["fromstring"]).fromstring(resp.content)
    id_el = root.find(f"{{{ATOM_NS}}}id")
    if id_el is not None:
        parts = id_el.text.split("/")
        list_id = parts[-1]
        print(f"[API] list created: {list_id} ({list_title})")
        return list_id
    return None

# ---------------------------------------------------------------------------
# Step 2 — Create Import Task
# ---------------------------------------------------------------------------


def create_import_task(list_id: str, task_title: str) -> str | None:
    """
    POST /contactimport/ to create an import task.
    Returns importId on success.
    """
    body = (
        f'<entry xmlns="{ATOM_NS}" xmlns:um="{UM_NS}">\n'
        f'  <title>{sax.escape(task_title)}</title>\n'
        f'  <um:type>UpdateExistsAddNew</um:type>\n'
        f'  <um:reportOpen>false</um:reportOpen>\n'
        f'  <um:importMethod>api</um:importMethod>\n'
        f'  <link href="{HOST}/list/{list_id}" rel="related"></link>\n'
        f'  <um:status>building</um:status>\n'
        f'</entry>'
    )

    resp = requests.post(
        f"{HOST}/contactimport/",
        params={"apikey": API_KEY, "method": "post", "alt": "atom"},
        headers=_api_headers(),
        auth=_api_auth(),
        data=body.encode("utf-8"),
        timeout=30,
    )

    if resp.status_code not in (200, 201):
        text = parse_xml_response(resp)
        print(f"[API] create import task failed: {resp.status_code} — {text}", file=sys.stderr)
        return None

    root = __import__("xml.etree.ElementTree", fromlist=["fromstring"]).fromstring(resp.content)
    id_el = root.find(f"{{{ATOM_NS}}}id")
    if id_el is not None:
        parts = id_el.text.split("/")
        import_id = parts[-1]
        print(f"[API] import task created: {import_id}")
        return import_id
    return None

# ---------------------------------------------------------------------------
# Step 3 — Submit Contacts
# ---------------------------------------------------------------------------


def submit_contacts(import_id: str, csv_path: str, attrs: list[tuple[str, str]]) -> int:
    """
    POST /contactimport/import/{importId}/ to submit contact entries from CSV.
    Returns the number of contacts submitted on success, 0 on failure.
    """
    # Build header→system_name mapping (case-insensitive)
    attr_map = {}
    for label, sys_name in attrs:
        attr_map[label.lower()] = sys_name

    # Read CSV data rows
    with open(csv_path, encoding="gbk", newline="") as f:
        reader = csv.reader(f)
        csv_header = next(reader)
        data_rows = list(reader)

    # Find Email column
    email_col = None
    for i, col in enumerate(csv_header):
        if col.strip().lower() == "email":
            email_col = i
            break

    if email_col is None:
        print("[API] no Email column found in CSV", file=sys.stderr)
        return 0

    # Build feed XML
    entries = []
    for row in data_rows:
        email_val = row[email_col].strip() if email_col < len(row) else ""
        if not email_val:
            continue
        entry_xml = f'  <entry xmlns:um="{UM_NS}">\n'
        entry_xml += f'    <email>{sax.escape(email_val)}</email>\n'
        # Only include attributes that have values (case-insensitive match)
        for i, col in enumerate(csv_header):
            key = col.strip().lower()
            if key in attr_map and i < len(row) and row[i].strip():
                entry_xml += f'    <um:attribute name="{attr_map[key]}">{sax.escape(row[i].strip())}</um:attribute>\n'
        entry_xml += f"  </entry>\n"
        entries.append(entry_xml)

    feed = (
        f'<feed xmlns="{ATOM_NS}">\n'
        f'  <link href="{HOST}/contactimport/{import_id}" rel="related"></link>\n'
        + "".join(entries)
        + f'</feed>'
    )

    # Scale timeout with contact count — 60s base + 2s per 10 contacts
    timeout = min(60 + len(entries) * 2, 300)

    resp = requests.post(
        f"{HOST}/contactimport/import/{import_id}/",
        params={"apikey": API_KEY, "method": "post", "alt": "atom"},
        headers=_api_headers(),
        auth=_api_auth(),
        data=feed.encode("utf-8"),
        timeout=timeout,
    )

    if resp.status_code not in (200, 201, 202):
        text = parse_xml_response(resp)
        print(f"[API] submit contacts failed: {resp.status_code} — {text}", file=sys.stderr)
        return 0

    root = __import__("xml.etree.ElementTree", fromlist=["fromstring"]).fromstring(resp.content)
    status_el = root.find(f"{{{UM_NS}}}status")
    total_el = root.find(f"{{{UM_NS}}}total")
    status = status_el.text if status_el is not None else "?"
    total = int(total_el.text) if total_el is not None else 0
    print(f"[API] contacts submitted: {total} contacts, status={status}")
    return total

# ---------------------------------------------------------------------------
# Step 4 — Execute Import Task
# ---------------------------------------------------------------------------


def execute_import(import_id: str) -> bool:
    """
    POST /contactimport/{importId}?method=put to execute the import task.
    Returns True on success.
    """
    body = (
        f'<entry xmlns="{ATOM_NS}" xmlns:um="{UM_NS}">\n'
        f'  <um:status>executing</um:status>\n'
        f'</entry>'
    )

    resp = requests.post(
        f"{HOST}/contactimport/{import_id}",
        params={"apikey": API_KEY, "method": "put", "alt": "atom"},
        headers=_api_headers(),
        auth=_api_auth(),
        data=body.encode("utf-8"),
        timeout=30,
    )

    if resp.status_code not in (200, 201, 202):
        text = parse_xml_response(resp)
        print(f"[API] execute import failed: {resp.status_code} — {text}", file=sys.stderr)
        return False

    return True

# ---------------------------------------------------------------------------
# Step 5 — Poll Import Result
# ---------------------------------------------------------------------------


def poll_import_status(import_id: str, contact_count: int = 0) -> dict:
    """
    Poll GET /contactimport/{importId}/ until final status.
    Auto-scales wait based on contact count: 60s base + 2s per 10 contacts, cap 300s.
    Returns result dict with status and counts.
    """
    max_wait = min(60 + contact_count * 2, 300)

    final_statuses = {"导入成功", "execute_succeed", "execute_failure", "execute_stop"}
    elapsed = 0

    while elapsed < max_wait:
        resp = requests.get(
            f"{HOST}/contactimport/{import_id}/",
            params={"apikey": API_KEY, "method": "get", "alt": "atom"},
            headers=_api_headers(),
            auth=_api_auth(),
            timeout=30,
        )

        if resp.status_code != 200:
            print(f"[API] poll failed: {resp.status_code}", file=sys.stderr)
            return {"status": "error"}

        root = __import__("xml.etree.ElementTree", fromlist=["fromstring"]).fromstring(resp.content)

        def get_text(tag):
            el = root.find(f"{{{UM_NS}}}{tag}")
            return el.text if el is not None else None

        status = get_text("status")
        total = get_text("total")
        valid = get_text("validNum")
        invalid = get_text("inValidNum")
        add_to_list = get_text("addToListSuccessNum")
        add_new = get_text("addSuccessNum")
        updated = get_text("updateSuccessNum")

        print(f"[POLL] status={status} total={total} valid={valid} invalid={invalid} added={add_to_list} new={add_new} updated={updated}")

        if status in final_statuses:
            return {
                "status": status,
                "total": total,
                "validNum": valid,
                "inValidNum": invalid,
                "addToListSuccessNum": add_to_list,
                "addSuccessNum": add_new,
                "updateSuccessNum": updated,
            }

        time.sleep(3)
        elapsed += 3

    return {"status": "timeout", "importId": import_id}

# ---------------------------------------------------------------------------
# Shared import pipeline
# ---------------------------------------------------------------------------


def import_csv_to_list(csv_path: str, list_title: str, attrs: list[tuple[str, str]]) -> dict:
    """
    Run the full Unimarketing import pipeline for a CSV file.
    Returns result dict from poll_import_status.
    """
    # Step 1: Create list
    list_id = create_list(list_title, attrs)
    if not list_id:
        print("Error: failed to create list", file=sys.stderr)
        sys.exit(1)

    # Step 2: Create import task
    task_title = f"API导入_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    import_id = create_import_task(list_id, task_title)
    if not import_id:
        print("Error: failed to create import task", file=sys.stderr)
        sys.exit(1)

    # Step 3: Submit contacts
    contact_count = submit_contacts(import_id, csv_path, attrs)
    if not contact_count:
        print("Error: failed to submit contacts", file=sys.stderr)
        sys.exit(1)

    # Step 4: Execute import
    if not execute_import(import_id):
        print("Error: failed to execute import", file=sys.stderr)
        sys.exit(1)

    # Step 5: Poll result
    print()
    result = poll_import_status(import_id, contact_count)

    print()
    if result["status"] in ("导入成功", "execute_succeed"):
        print(f"SUCCESS: Import complete — listId={list_id}, importId={import_id}")
        print(f"  Total: {result.get('total')} | Valid: {result.get('validNum')} | "
              f"Invalid: {result.get('inValidNum')} | Added: {result.get('addToListSuccessNum')} | "
              f"New: {result.get('addSuccessNum')} | Updated: {result.get('updateSuccessNum')}")
    else:
        print(f"FAILED: Import ended with status={result['status']}")
        sys.exit(1)

    return result

# ---------------------------------------------------------------------------
# Main runners
# ---------------------------------------------------------------------------


def _extract_sn(xlsx_path: str) -> str | None:
    """Try to find SN-xxxxx in the xlsx file path."""
    sn_match = re.search(r"SN-\d+", os.path.basename(xlsx_path))
    if not sn_match:
        sn_match = re.search(r"SN-\d+", xlsx_path)
    return sn_match.group(0) if sn_match else None


def run_test(xlsx_path: str, sn: str | None = None):
    """Test mode: generate test CSV and import."""
    config = _load_config()
    test_emails = config.get("test_emails", [])
    if not test_emails:
        print("Error: no test_emails in config.json", file=sys.stderr)
        sys.exit(1)

    output_dir = os.path.dirname(xlsx_path)

    if not sn:
        sn = _extract_sn(xlsx_path)
    if not sn:
        print("Error: cannot find SN number. Provide --sn explicitly.", file=sys.stderr)
        sys.exit(1)

    xlsx_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    now = datetime.now().strftime("%Y%m%d%H%M%S")
    list_title = f"test_{sn}_{xlsx_name}_{now}"

    print(f"=== Unimarketing Test List Import ===")
    print(f"SN: {sn}")
    print(f"List: {list_title}")
    print(f"Test emails: {test_emails}")
    print()

    csv_path = generate_test_csv(xlsx_path, output_dir, test_emails)

    # Read CSV header for attribute detection
    with open(csv_path, encoding="gbk", newline="") as f:
        csv_header = next(csv.reader(f))

    attrs = get_attr_mapping(csv_header)
    if not attrs:
        print("[WARN] No Token/SubId columns found in CSV header", file=sys.stderr)
    print(f"Attributes: {attrs}")
    print()

    import_csv_to_list(csv_path, list_title, attrs)


def run_formal(xlsx_path: str, sn: str | None = None):
    """Formal mode: generate formal CSV (all rows) and import."""
    output_dir = os.path.dirname(xlsx_path)

    if not sn:
        sn = _extract_sn(xlsx_path)
    if not sn:
        print("Error: cannot find SN number. Provide --sn explicitly.", file=sys.stderr)
        sys.exit(1)

    xlsx_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    now = datetime.now().strftime("%Y%m%d%H%M%S")
    list_title = f"formal_{sn}_{xlsx_name}_{now}"

    print(f"=== Unimarketing Formal List Import ===")
    print(f"SN: {sn}")
    print(f"List: {list_title}")
    print()

    csv_path = generate_formal_csv(xlsx_path, output_dir)

    # Read CSV header for attribute detection
    with open(csv_path, encoding="gbk", newline="") as f:
        csv_header = next(csv.reader(f))

    attrs = get_attr_mapping(csv_header)
    if not attrs:
        print("[WARN] No Token/SubId columns found in CSV header", file=sys.stderr)
    print(f"Attributes: {attrs}")
    print()

    import_csv_to_list(csv_path, list_title, attrs)


def main():
    parser = argparse.ArgumentParser(
        description="Unimarketing Contact List Import — generate CSV from xlsx and import to a new list"
    )
    parser.add_argument("--xlsx", required=True, help="Path to xlsx contact file")
    parser.add_argument("--sn", default=None, help="SN number (auto-detected if not provided)")
    parser.add_argument("--formal", action="store_true", help="Import all rows as formal list (default: test mode)")
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"Error: file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    if args.formal:
        run_formal(args.xlsx, args.sn)
    else:
        run_test(args.xlsx, args.sn)


if __name__ == "__main__":
    main()
