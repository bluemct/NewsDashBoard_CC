"""
Convert .xlsx files to CSV files with full cell verification.

Reads every cell from the source xlsx, writes a comma-delimited CSV,
then re-reads the CSV and compares every cell value against the original
to catch any encoding or formatting loss.
"""
import argparse
import csv
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def read_xlsx(path):
    """Read all sheets from an xlsx file, return {sheet_name: [(row_values), ...]}"""
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[name] = rows
    wb.close()
    return sheets


def normalize_cell(val):
    """Normalize a cell value to a comparable string."""
    if val is None:
        return ""
    return str(val).strip()


def xlsx_to_csv(xlsx_path, csv_path, encoding="gb18030"):
    """Convert first sheet of xlsx to a CSV file with specified encoding."""
    sheets = read_xlsx(xlsx_path)
    sheet_name = list(sheets.keys())[0]
    rows = sheets[sheet_name]

    if not rows:
        print(f"Warning: sheet '{sheet_name}' is empty.")
        with open(csv_path, "w", encoding=encoding, errors="replace", newline="") as f:
            pass
        return True

    with open(csv_path, "w", encoding=encoding, errors="replace", newline="") as f:
        writer = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            str_row = [""] if row == [None] else [str(c) if c is not None else "" for c in row]
            writer.writerow(str_row)

    return True


def verify_conversion(xlsx_path, csv_path, encoding="gb18030"):
    """Re-read CSV and compare every cell against the original xlsx."""
    sheets = read_xlsx(xlsx_path)
    sheet_name = list(sheets.keys())[0]
    xlsx_rows = sheets[sheet_name]

    # Read CSV back
    csv_rows = []
    with open(csv_path, "r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=",")
        for row in reader:
            csv_rows.append(row)

    # Compare
    mismatches = []
    max_rows = max(len(xlsx_rows), len(csv_rows))

    for r in range(max_rows):
        if r >= len(xlsx_rows):
            for c, val in enumerate(csv_rows[r]):
                col_letter = _idx_to_col(c)
                mismatches.append(
                    f"  [{sheet_name}] {col_letter}{r + 1}: "
                    f"extra row in CSV, value='{val}'"
                )
            continue

        if r >= len(csv_rows):
            mismatches.append(
                f"  [{sheet_name}] row {r + 1}: missing from CSV"
            )
            continue

        xlsx_row = xlsx_rows[r]
        csv_row = csv_rows[r]
        max_cols = max(len(xlsx_row), len(csv_row))

        for c in range(max_cols):
            col_letter = _idx_to_col(c)
            cell_ref = f"{col_letter}{r + 1}"

            if c >= len(xlsx_row):
                mismatches.append(
                    f"  [{sheet_name}] {cell_ref}: "
                    f"extra column in CSV, value='{csv_row[c]}'"
                )
                continue

            if c >= len(csv_row):
                orig = normalize_cell(xlsx_row[c])
                if orig:
                    mismatches.append(
                        f"  [{sheet_name}] {cell_ref}: "
                        f"missing from CSV (original='{orig}')"
                    )
                continue

            orig_val = normalize_cell(xlsx_row[c])
            csv_val = csv_row[c].strip()

            if orig_val != csv_val:
                mismatches.append(
                    f"  [{sheet_name}] {cell_ref}: "
                    f"original='{orig_val}' vs csv='{csv_val}'"
                )

    return mismatches


def _idx_to_col(idx):
    """Convert 0-based column index to Excel column letter (A, B, ..., Z, AA, ...)."""
    letters = ""
    idx += 1  # 1-based
    while idx > 0:
        idx, remainder = divmod(idx - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def main():
    parser = argparse.ArgumentParser(
        description="Convert .xlsx to CSV with cell-by-cell verification."
    )
    parser.add_argument("input", help="Path to source .xlsx file")
    parser.add_argument(
        "output", nargs="?", default=None,
        help="Path to output CSV file (default: same name with .csv extension)"
    )
    parser.add_argument(
        "--encoding", default="gb18030", choices=["gb18030", "gb2312", "gbk", "utf-8"],
        help="Output encoding (default: gb18030)"
    )
    args = parser.parse_args()

    xlsx_path = args.input
    if not os.path.isfile(xlsx_path):
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    csv_path = args.output if args.output else os.path.splitext(xlsx_path)[0] + ".csv"
    encoding = args.encoding

    # Auto-fallback chain: try requested encoding first, then wider encodings
    encodings_to_try = [encoding]
    if encoding not in ("gb18030", "utf-8"):
        encodings_to_try.append("gb18030")
    encodings_to_try.append("utf-8")

    used_encoding = encodings_to_try[0]
    for enc in encodings_to_try:
        try:
            xlsx_to_csv(xlsx_path, csv_path, enc)
            used_encoding = enc
            break
        except UnicodeEncodeError:
            continue
    else:
        print("Error: all encodings failed.", file=sys.stderr)
        sys.exit(1)

    print(f"Input:  {xlsx_path}")
    print(f"Output: {csv_path}")

    if used_encoding != encoding:
        print(f"Note: {encoding} failed, fell back to {used_encoding}")
    else:
        print(f"Encoding: {encoding}")

    print(f"  CSV written to {csv_path}")

    # Step 2: Verify
    print(f"\n[2/2] Verifying every cell against original (encoding={used_encoding})...")
    mismatches = verify_conversion(xlsx_path, csv_path, used_encoding)

    if not mismatches:
        print("  All cells match. Conversion verified successfully.")
        sys.exit(0)
    else:
        total = len(mismatches)
        print(f"\n  Found {total} mismatch(es):")
        for m in mismatches:
            print(m)
        sys.exit(1)


if __name__ == "__main__":
    main()
