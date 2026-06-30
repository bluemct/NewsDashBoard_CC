"""
EDM Email Processor — Windows GUI Tool

Standalone tkinter application for processing EDM emails:
- Extract SN from .msg, create folder in output directory
- Extract nested .msg (0 recipients), convert to HTML with token replacement
- Generate formal and test CSVs
- Log all steps to process.log

Output (exactly 5 files):
  1. Nested .msg
  2. EDM_template.html
  3. formal_*.csv
  4. test_*.csv
  5. process.log

Usage:
    python edm_gui.py

Config files (same directory as this script/exe):
    config.json        — test email addresses
    Tokenmapping.json  — token name→value mappings
"""
import csv
import glob
import json
import os
import re
import shutil
import sys
import threading
from urllib import parse as urllib_parse
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

# ---------------------------------------------------------------------------
# Resolve script directory — works for both .py and PyInstaller .exe
# ---------------------------------------------------------------------------
if getattr(sys, "frozen", False):
    # Running as PyInstaller exe — config files are next to the exe
    _SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    _SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Import core functions from edm_process
# ---------------------------------------------------------------------------
_SKILL_DIR = os.path.join(_SCRIPT_DIR, ".claude", "skills", "edm-process")
if not getattr(sys, "frozen", False) and _SKILL_DIR not in sys.path:
    sys.path.insert(0, _SKILL_DIR)

try:
    from edm_process import (
        _get_short_path,
        extract_sn,
        find_target_attachment_idx,
        replace_span_tokens,
        save_target_attachment,
    )
except ImportError as e:
    print(f"Warning: Could not import edm_process: {e}", file=sys.stderr)

# ---------------------------------------------------------------------------
# Import xlsx_to_csv functions directly (no subprocess needed)
# ---------------------------------------------------------------------------
_XLSX_SKILL_DIR = os.path.join(_SCRIPT_DIR, ".claude", "skills", "xlsx-to-csv")
if not getattr(sys, "frozen", False) and _XLSX_SKILL_DIR not in sys.path:
    sys.path.insert(0, _XLSX_SKILL_DIR)

try:
    from xlsx_to_csv import xlsx_to_csv as _xlsx_to_csv
except ImportError as e:
    print(f"Warning: Could not import xlsx_to_csv: {e}", file=sys.stderr)

# ---------------------------------------------------------------------------
# Import Unimarketing contact import functions
# ---------------------------------------------------------------------------
_UM_SKILL_DIR = os.path.join(_SCRIPT_DIR, ".claude", "skills", "unimarketing-contactimport2list")
_EDM_BUILD_DIR = os.path.join(_SCRIPT_DIR, "_edm_build")

if not getattr(sys, "frozen", False) and _UM_SKILL_DIR not in sys.path:
    sys.path.insert(0, _UM_SKILL_DIR)
elif getattr(sys, "frozen", False) and _EDM_BUILD_DIR not in sys.path:
    sys.path.insert(0, _EDM_BUILD_DIR)

try:
    from unimarketing_test_list import (
        generate_test_csv as _um_generate_test_csv,
        generate_formal_csv as _um_generate_formal_csv,
        get_attr_mapping as _um_get_attr_mapping,
        create_list as _um_create_list,
        create_import_task as _um_create_import_task,
        submit_contacts as _um_submit_contacts,
        execute_import as _um_execute_import,
        poll_import_status as _um_poll_import_status,
    )
except ImportError as e:
    print(f"Warning: Could not import unimarketing_test_list: {e}", file=sys.stderr)

# ---------------------------------------------------------------------------
# Import verify_list_contacts for post-import verification
# ---------------------------------------------------------------------------
def _find_verify_module(name: str) -> str | None:
    """Find a module file — tries script dir first, then _internal/ (exe mode)."""
    direct = os.path.join(_SCRIPT_DIR, name)
    if os.path.isfile(direct):
        return direct
    internal = os.path.join(_SCRIPT_DIR, "_internal", name)
    if os.path.isfile(internal):
        return internal
    return None

_verify_path = _find_verify_module("verify_list_contacts.py")
if _verify_path:
    try:
        import importlib.util
        _spec = importlib.util.spec_from_file_location("verify_list_contacts", _verify_path)
        _verify_mod = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(_verify_mod)
        _verify_list_import = _verify_mod.verify_list_import
        _verify_find_lists_by_sn = _verify_mod.find_lists_by_sn
        _verify_read_xlsx_emails = _verify_mod.read_xlsx_emails
        _verify_get_list_info = _verify_mod.get_list_info
        _HAS_VERIFY = True
    except Exception as e:
        print(f"Warning: Could not load verify_list_contacts: {e}", file=sys.stderr)
        _HAS_VERIFY = False
else:
    _HAS_VERIFY = False

# ---------------------------------------------------------------------------
# Import deep_verify for field-by-field comparison
# ---------------------------------------------------------------------------
_deep_verify_path = _find_verify_module("deep_verify_list.py")
if _deep_verify_path:
    try:
        import importlib.util
        _dv_spec = importlib.util.spec_from_file_location("deep_verify_list", _deep_verify_path)
        _dv_mod = importlib.util.module_from_spec(_dv_spec)
        _dv_spec.loader.exec_module(_dv_mod)
        _deep_verify = _dv_mod.deep_verify
        _HAS_DEEP_VERIFY = True
    except Exception as e:
        print(f"Warning: Could not load deep_verify_list: {e}", file=sys.stderr)
        _HAS_DEEP_VERIFY = False
else:
    _HAS_DEEP_VERIFY = False


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_OUTPUT_BASE = os.path.join(os.path.expanduser("~"), "Desktop", "EDM")

CONFIG_PATH = os.path.join(_SCRIPT_DIR, "config.json")
TOKENMAP_PATH = os.path.join(_SCRIPT_DIR, "Tokenmapping.json")
XLSX_SEARCH_DIR_CONFIG = os.path.join(_SCRIPT_DIR, "xlsx_search_dir.json")

DEFAULT_XLSX_SEARCH_DIR = r"C:\Users\SI-Agent\AgentProject\Microsoft\Azure Service Notifications Collaboration - 2026"

# PyInstaller puts datas in _internal/ next to the exe
def _find_config(name: str) -> str:
    """Find a config file — tries exe dir first, then _internal/."""
    direct = os.path.join(_SCRIPT_DIR, name)
    if os.path.isfile(direct):
        return direct
    internal = os.path.join(_SCRIPT_DIR, "_internal", name)
    if os.path.isfile(internal):
        return internal
    return direct  # fall back to direct even if missing

def _resolve_config(path: str) -> str:
    """Resolve CONFIG_PATH or TOKENMAP_PATH to actual file location."""
    return _find_config(os.path.basename(path))


def _xlsx_config_path() -> str:
    """Return the actual path for reading/writing xlsx_search_dir.json."""
    if getattr(sys, "frozen", False):
        # PyInstaller: read/write from _internal/
        return os.path.join(_SCRIPT_DIR, "_internal", "xlsx_search_dir.json")
    # Dev mode: project root
    return XLSX_SEARCH_DIR_CONFIG


def _load_xlsx_search_dir() -> str:
    """Load xlsx_search_dir.json — returns default path if missing."""
    p = _xlsx_config_path()
    if os.path.isfile(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f).get("search_directory", DEFAULT_XLSX_SEARCH_DIR)
        except (json.JSONDecodeError, KeyError):
            pass
    return DEFAULT_XLSX_SEARCH_DIR


def _save_xlsx_search_dir(path: str) -> None:
    """Save search_directory to xlsx_search_dir.json."""
    p = _xlsx_config_path()
    os.makedirs(os.path.dirname(p), exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        json.dump({"search_directory": path}, f, indent=2, ensure_ascii=False)
        f.write("\n")


def extract_xlsx_filename_from_msg(msg_path: str) -> str | None:
    """Read .msg body text, extract SharePoint URL, and return the xlsx filename.

    Example URL:
        https://microsoftapc.sharepoint.com/.../Token%201-13%20SN-56699.xlsx?d=...
    Returns:
        Token 1-13 SN-56699.xlsx
    """
    try:
        from extract_msg import Message as MsgParser
        msg = MsgParser(msg_path)
        body = msg.body or ""
        msg.close()
    except Exception:
        return None

    # Find SharePoint-like URLs containing .xlsx
    urls = re.findall(r'https?://[^\s<>"\']+\.xlsx[^\s<>"\']*', body)
    if not urls:
        return None

    url = urls[0]
    # Get the path portion: take text between last / and first ?
    after_last_slash = url.rsplit("/", 1)[-1]
    filename = after_last_slash.split("?")[0]
    # URL-decode the filename
    filename = urllib_parse.unquote(filename)
    if not filename:
        return None
    return filename


def discover_xlsx(sn: str, search_dir: str, filename_hint: str | None = None) -> str | None:
    """Recursively search search_dir for an .xlsx file inside a folder matching sn.

    Searches for folders whose name contains the SN (e.g. 'SN-53672' or 'SN 53672')
    and returns the first .xlsx found inside.

    If *filename_hint* is provided, the function first does a full recursive scan
    of search_dir for an exact (case-insensitive) filename match — regardless of
    folder naming.  If that fails (or filename_hint is None), it falls back to
    the SN-folder-based search.
    """
    if not os.path.isdir(search_dir):
        return None

    # Priority 1: global filename exact match (ignores folder naming)
    if filename_hint:
        hint_lower = filename_hint.lower()
        for root, dirs, files in os.walk(search_dir):
            for f in files:
                if f.lower() == hint_lower:
                    return os.path.join(root, f)

    # Priority 2: SN folder match (original logic) — skip if SN is a dummy
    if sn is None or sn == "SN":
        return None
    sn_no_dash = sn.replace("-", "")  # SN-53672 → SN53672

    for root, dirs, files in os.walk(search_dir):
        folder_name = os.path.basename(root).lower()
        # Match folder names like "SN-53672", "SN 53672", "SN53672", or any containing the digits
        if sn_no_dash.lower() in folder_name.replace("-", "") or sn.lower() in folder_name:
            xlsx_files = [f for f in files if f.lower().endswith(".xlsx")]
            if not xlsx_files:
                continue
            return os.path.join(root, xlsx_files[0])

    return None

# ---------------------------------------------------------------------------
# Config file loading
# ---------------------------------------------------------------------------

def _load_config() -> dict:
    """Load config.json from script directory. Returns defaults if missing."""
    defaults = {
        "test_emails": [
            "ma.chuntao@oe.21vianet.com",
            "microsoft.163163@163.com",
        ]
    }
    _path = _resolve_config(CONFIG_PATH)
    if os.path.isfile(_path):
        with open(_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        defaults.update(data)
    return defaults

def _load_tokenmap() -> dict:
    """Load Tokenmapping.json from script directory. Returns empty dict if missing."""
    _path = _resolve_config(TOKENMAP_PATH)
    if not os.path.isfile(_path):
        return {}
    with open(_path, "r", encoding="utf-8") as f:
        mapping_list = json.load(f)
    return {item["Name"]: item["Value"] for item in mapping_list}

def _load_raw_json(path: str) -> str | None:
    """Return raw file content or None."""
    if not os.path.isfile(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

# ---------------------------------------------------------------------------
# ProcessLogger — logs to GUI + in-memory list (written to file at end)
# ---------------------------------------------------------------------------


def _log_dir() -> str:
    """Return the central log directory path.
    Frozen (PyInstaller): exe_dir/_internal/Log/
    Dev mode: project_root/Log/
    """
    if getattr(sys, "frozen", False):
        return os.path.join(_SCRIPT_DIR, "_internal", "Log")
    return os.path.join(_SCRIPT_DIR, "Log")


def _ensure_log_dir() -> str:
    """Ensure log directory exists and return its path."""
    d = _log_dir()
    os.makedirs(d, exist_ok=True)
    return d


class ProcessLogger:
    def __init__(self, root: tk.Tk, text_widget: tk.Text, log_file: str = None):
        self.root = root
        self.text_widget = text_widget
        self.lines: list[str] = []
        self.log_file = log_file

    def log(self, message: str):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{timestamp}] {message}"
        self.lines.append(line)
        self.root.after(0, self._append_gui, line)
        # Also write to log file immediately
        if self.log_file:
            try:
                with open(self.log_file, "a", encoding="utf-8") as f:
                    f.write(line + "\n")
            except OSError as e:
                print(f"[LOG] failed to write to {self.log_file}: {e}", file=sys.stderr)

    def _append_gui(self, line: str):
        self.text_widget.config(state="normal")
        self.text_widget.insert("end", line + "\n")
        self.text_widget.see("end")
        self.text_widget.config(state="disabled")

    def write_file(self, path: str):
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(self.lines) + "\n")


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------


def _convert_xlsx_to_csv(xlsx_path: str, logger: ProcessLogger) -> bool:
    """Convert xlsx to CSV using openpyxl directly (no subprocess)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.log("[CSV] openpyxl not available")
        return False

    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        csv_path = os.path.splitext(xlsx_path)[0] + ".csv"
        encodings_to_try = ["gb18030", "utf-8"]
        success = False

        try:
            for enc in encodings_to_try:
                try:
                    with open(csv_path, "w", encoding=enc, errors="replace", newline="") as f:
                        writer = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL)
                        for row in ws.iter_rows(values_only=True):
                            str_row = [""] if row == [None] else [str(c) if c is not None else "" for c in row]
                            writer.writerow(str_row)
                    success = True
                    logger.log(f"[CSV] {os.path.basename(csv_path)} ({enc})")
                    break
                except UnicodeEncodeError:
                    continue

            if not success:
                logger.log("[CSV] all encodings failed")
        finally:
            wb.close()

        return success
    except Exception as e:
        logger.log(f"[CSV] xlsx_to_csv failed: {e}")
        return False


def _generate_formal_test_csv(xlsx_path: str, config: dict, logger: ProcessLogger) -> None:
    """Generate formal_*.csv (all rows) and test_*.csv (2 rows) from the source CSV."""
    sn_dir = os.path.dirname(xlsx_path)
    base = os.path.splitext(os.path.basename(xlsx_path))[0]

    csv_files = glob.glob(os.path.join(sn_dir, base + "*.csv"))
    csv_path = csv_files[0] if csv_files else None
    if not csv_path or not os.path.exists(csv_path):
        logger.log("[CSV] no source CSV found for formal/test generation")
        return

    with open(csv_path, encoding="gb18030", newline="") as f:
        reader = list(csv.reader(f))
        header = reader[0]
        rows = reader[1:]

    # Formal CSV — copy all rows
    formal_path = os.path.join(sn_dir, f"formal_{base}.csv")
    shutil.copy2(csv_path, formal_path)
    logger.log(f"[CSV-FORMAL] saved: {os.path.basename(formal_path)} ({len(rows)} rows)")

    # Test CSV — pick rows with most tokens filled, replace email column
    test_emails = config.get("test_emails", [
        "ma.chuntao@oe.21vianet.com",
        "microsoft.163163@163.com",
    ])
    test_count = max(len(test_emails), 2)

    email_idx = None
    for i, col in enumerate(header):
        if col.strip().lower() == "email":
            email_idx = i
            break

    token_cols = [i for i, col in enumerate(header) if col.strip().lower().startswith("token")]
    row_scores = []
    for i, row in enumerate(rows):
        score = sum(1 for idx in token_cols if idx < len(row) and row[idx].strip())
        if score > 0:
            row_scores.append((score, i, row))
    row_scores.sort(reverse=True)

    # Collect up to test_count distinct rows (round-robin from top-scored)
    selected = []
    idx = 0
    while len(selected) < test_count:
        if idx < len(row_scores):
            selected.append(list(row_scores[idx][2]))
        elif len(rows) > len(selected):
            selected.append(list(rows[len(selected)]))
        else:
            selected.append(["" for _ in header])
        idx += 1

    if email_idx is not None:
        for i, row in enumerate(selected):
            if i < len(test_emails):
                row[email_idx] = test_emails[i]

    test_path = os.path.join(sn_dir, f"test_{base}.csv")
    with open(test_path, "w", encoding="gb18030", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in selected:
            writer.writerow(row)
    logger.log(f"[CSV-TEST] saved: {os.path.basename(test_path)} ({len(selected)} rows)")

    # Clean up raw CSV — only keep formal_ and test_
    os.remove(csv_path)
    logger.log(f"[CLEANUP] removed raw CSV: {os.path.basename(csv_path)}")


def _convert_msg_to_html(
    msg_path: str,
    output_html: str,
    token_mapping: dict | None,
    logger: ProcessLogger,
) -> bool:
    """Convert .msg to HTML via win32com, with token replacement."""
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
    except Exception as e:
        logger.log(f"[HTML] could not connect to Outlook: {e}")
        pythoncom.CoUninitialize()
        return False

    short_path = _get_short_path(msg_path)
    try:
        msg = namespace.OpenSharedItem(short_path)
    except Exception as e:
        logger.log(f"[HTML] could not open .msg: {e}")
        pythoncom.CoUninitialize()
        return False

    html_body = msg.HTMLBody or ""
    subject = msg.Subject or ""

    try:
        msg.Close(0)
    except Exception:
        pass

    pythoncom.CoUninitialize()

    if not html_body:
        logger.log("[HTML] no HTMLBody found")
        return False

    # Insert subject line
    safe_subject = subject.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    subject_block = (
        f"<p class=MsoNormal><b><span lang=ZH-CN\n"
        f"style='font-family:等线;mso-hansi-font-family:Calibri;mso-bidi-font-family:等线;\n"
        f"color:black'>主题</span></b><span style='font-family:等线;mso-hansi-font-family:Calibri;\n"
        f"mso-bidi-font-family:等线;color:black'>: {safe_subject}</span></p>\n"
        f"\n"
        f"<p class=MsoNormal><o:p>&nbsp;</o:p></p>\n"
    )
    body_pos = html_body.lower().find("<body")
    if body_pos > 0:
        body_close = html_body.find(">", body_pos)
        if body_close > 0:
            html_body = html_body[:body_close + 1] + "\n" + subject_block + html_body[body_close + 1:]

    # Token replacement
    if token_mapping:
        html_body = replace_span_tokens(html_body, token_mapping)

    # Remove _MailOriginal anchor, add blank line
    html_body = re.sub(
        r'<a\s+name="_MailOriginal">\s*<span[^>]*>\s*<o:p>\s*&nbsp;\s*</o:p>\s*</span>\s*</a>\s*</p>',
        "<p class=MsoNormal><o:p>&nbsp;</o:p></p>",
        html_body,
        flags=re.DOTALL | re.IGNORECASE,
    )

    # Fix charset — Outlook declares "unicode" (UTF-16LE) but we save as UTF-8
    html_body = html_body.replace(
        'meta http-equiv=Content-Type content="text/html; charset=unicode"',
        'meta http-equiv=Content-Type content="text/html; charset=utf-8"',
    )

    with open(output_html, "wb") as f:
        f.write(html_body.encode("utf-8"))

    size_kb = os.path.getsize(output_html) / 1024
    logger.log(f"[HTML] saved: {os.path.basename(output_html)} ({size_kb:.1f} KB)")
    return True


# ---------------------------------------------------------------------------
# Main processing thread
# ---------------------------------------------------------------------------


def _process(
    logger: ProcessLogger,
    msg_path: str,
    xlsx_path: str,
    output_base: str,
    config: dict,
    on_done: callable,
    on_error: callable,
):
    try:
        from extract_msg import Message as MsgParser
    except ImportError:
        on_error("Missing dependency 'extract-msg'. Run: pip install extract-msg")
        return

    logger.log("=== EDM Processing Started ===")
    logger.log(f"MSG file: {msg_path}")
    logger.log(f"XLSX file: {xlsx_path}")

    # Step 1: Extract SN
    msg = MsgParser(msg_path)
    subject = msg.subject or ""
    logger.log(f"Subject: {subject}")

    sn = extract_sn(subject)
    if not sn:
        sn = extract_sn(msg_path)
    if not sn:
        msg.close()
        on_error("No SN number found. Subject or filename should contain SN-xxxxx.")
        return

    logger.log(f"SN extracted: {sn}")

    # Step 2: Create folder
    sn_folder = os.path.join(output_base, sn)
    os.makedirs(sn_folder, exist_ok=True)
    logger.log(f"Output folder: {sn_folder}")

    # Step 3: Copy xlsx & convert to CSV
    xlsx_dst = os.path.join(sn_folder, os.path.basename(xlsx_path))
    shutil.copy2(xlsx_path, xlsx_dst)
    logger.log(f"[COPY] {os.path.basename(xlsx_path)} -> {sn}/")
    _convert_xlsx_to_csv(xlsx_dst, logger)

    # Step 4: Generate formal & test CSVs (removes raw CSV)
    _generate_formal_test_csv(xlsx_dst, config, logger)

    # Step 5: Extract nested .msg
    target_idx = find_target_attachment_idx(msg_path)
    attach_path = None
    if target_idx is not None and target_idx < len(msg.attachments):
        matched_att = msg.attachments[target_idx]
        attach_path = save_target_attachment(matched_att, sn_folder)
        logger.log(f"[ATTACH] saved nested .msg")
    else:
        logger.log("[ATTACH] no nested .msg found, skipping HTML conversion")
    msg.close()

    # Step 6: Convert to HTML
    if attach_path:
        logger.log("[HTML] converting to HTML via Outlook...")
        html_path = os.path.join(sn_folder, "EDM_template.html")

        token_mapping = _load_tokenmap()
        if token_mapping:
            logger.log(f"[TOKEN] loaded {len(token_mapping)} tokens from Tokenmapping.json")
        else:
            logger.log("[TOKEN] Tokenmapping.json not found, skipping token replacement")

        success = _convert_msg_to_html(attach_path, html_path, token_mapping, logger)
        if not success:
            logger.log("[HTML] conversion failed")

    # Step 7: Done
    logger.log("=== Processing Complete ===")
    logger.log(f"Output folder: {sn_folder}")

    on_done(sn_folder)


# ---------------------------------------------------------------------------
# Config Editor Dialog
# ---------------------------------------------------------------------------

class ConfigEditorDialog:
    """Simple modal dialog to view/edit a JSON file."""

    def __init__(self, parent: tk.Tk, title: str, path: str, content: str | None):
        self.path = path
        self.result = None

        self.win = tk.Toplevel(parent)
        self.win.title(title)
        self.win.geometry("500x400")
        self.win.resizable(True, True)
        self.win.transient(parent)
        self.win.grab_set()

        # Center on parent
        self.win.focus_set()

        # Top bar — title + buttons
        top = ttk.Frame(self.win, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text=f"{os.path.basename(path)}").pack(side="left", anchor="w", expand=True)

        if content is None:
            ttk.Label(top, text="File not found — creating new.", foreground="#e67e22").pack(side="left", padx=(0, 8))

        ttk.Button(top, text="Save", command=lambda: self._save(text)).pack(side="right", padx=(0, 6))
        ttk.Button(top, text="Cancel", command=self.win.destroy).pack(side="right")

        # Text area with scrollbar
        text_frame = ttk.Frame(self.win)
        text_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        sb = ttk.Scrollbar(text_frame, orient="vertical")
        sb.pack(side="right", fill="y")

        text = tk.Text(text_frame, font=("Consolas", 10), wrap="word",
                       padx=6, pady=6, yscrollcommand=sb.set)
        text.pack(fill="both", expand=True)
        sb.config(command=text.yview)

        if content:
            text.insert("1.0", content)
        else:
            # Default template based on file name
            if "tokenmap" in os.path.basename(path).lower():
                text.insert("1.0", "[\n  {\n    \"Name\": \"\",\n    \"Value\": \"\"\n  }\n]\n")
            else:
                text.insert("1.0", "{\n  \"test_emails\": [\n    \"\",\n    \"\"\n  ]\n}\n")

    def _save(self, text_widget: tk.Text):
        raw = text_widget.get("1.0", "end-1c")
        try:
            # Validate JSON
            parsed = json.loads(raw)
            # Re-format for consistency
            formatted = json.dumps(parsed, indent=2, ensure_ascii=False)
            with open(self.path, "w", encoding="utf-8") as f:
                f.write(formatted + "\n")
            self.result = parsed
        except json.JSONDecodeError as e:
            messagebox.showerror("Invalid JSON", str(e), parent=self.win)
            return
        self.win.destroy()


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------


class EDMGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("EDM Email Processor")
        self.root.geometry("800x600")
        self.root.minsize(640, 450)

        # State
        self.msg_var = tk.StringVar()
        self.xlsx_var = tk.StringVar()
        self.output_var = tk.StringVar(value=DEFAULT_OUTPUT_BASE)
        self.last_sn_folder = None
        self._menu_open_folder_cmd = None  # callback to enable menu item

        self._build_ui()
        self._build_menu()

    def _build_menu(self):
        """Build the top menu bar."""
        self._menubar = tk.Menu(self.root, font=("Microsoft YaHei UI", 9))
        self.root.config(menu=self._menubar)

        # File menu
        self._file_menu = tk.Menu(self._menubar, tearoff=0, font=("Microsoft YaHei UI", 9))
        self._menubar.add_cascade(label="文件", menu=self._file_menu)
        self._file_menu.add_command(label="打开输出文件夹", command=self._open_folder, accelerator="Ctrl+O", state="disabled")
        self._file_menu.add_command(label="修改输出目录...", command=self._browse_output)
        self._file_menu.add_separator()
        self._file_menu.add_command(label="退出", command=self.root.quit, accelerator="Alt+F4")

        # Settings menu
        settings_menu = tk.Menu(self._menubar, tearoff=0, font=("Microsoft YaHei UI", 9))
        self._menubar.add_cascade(label="设置", menu=settings_menu)
        settings_menu.add_command(label="编辑测试邮箱 (config.json)", command=lambda: self._menu_edit_config("config.json"), accelerator="Ctrl+E")
        settings_menu.add_command(label="编辑令牌映射 (Tokenmapping.json)", command=lambda: self._menu_edit_config("Tokenmapping.json"), accelerator="Ctrl+T")
        settings_menu.add_separator()
        settings_menu.add_command(label="XLSX 检索目录...", command=self._menu_edit_search_dir)

        # Help menu
        help_menu = tk.Menu(self._menubar, tearoff=0, font=("Microsoft YaHei UI", 9))
        self._menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="关于", command=self._show_about)

        # Keyboard shortcuts
        self.root.bind("<Control-o>", lambda e: self._open_folder())
        self.root.bind("<Control-e>", lambda e: self._menu_edit_config("config.json"))
        self.root.bind("<Control-t>", lambda e: self._menu_edit_config("Tokenmapping.json"))

    def _menu_edit_config(self, label: str):
        """Edit a config JSON file from the menu."""
        path = _resolve_config(CONFIG_PATH if label == "config.json" else TOKENMAP_PATH)
        content = _load_raw_json(path)
        dialog = ConfigEditorDialog(self.root, f"编辑 {os.path.basename(path)}", path, content)
        self.root.wait_window(dialog.win)

    def _menu_edit_search_dir(self):
        """Dialog to edit the XLSX search directory."""
        current = _load_xlsx_search_dir()
        dir_dialog = tk.Toplevel(self.root)
        dir_dialog.title("XLSX 检索目录")
        dir_dialog.geometry("550x120")
        dir_dialog.resizable(True, False)
        dir_dialog.transient(self.root)
        dir_dialog.grab_set()
        dir_dialog.focus_set()

        ttk.Label(dir_dialog, text="XLSX 文件检索目录路径：", font=("Microsoft YaHei UI", 9)).pack(
            anchor="w", padx=15, pady=(12, 4)
        )

        dir_var = tk.StringVar(value=current)
        entry_frame = ttk.Frame(dir_dialog, padding=(15, 0))
        entry_frame.pack(fill="x", pady=(0, 8))

        ttk.Entry(entry_frame, textvariable=dir_var, width=55, font=("Consolas", 9)).pack(
            side="left", fill="x", expand=True, padx=(0, 8)
        )
        ttk.Button(entry_frame, text="浏览...", command=lambda: (
            dir_var.set(filedialog.askdirectory(title="选择检索目录"))
        )).pack(side="right")

        result = [None]

        def save():
            val = dir_var.get().strip()
            if val:
                _save_xlsx_search_dir(val)
                result[0] = val
            dir_dialog.destroy()

        btn_frame = ttk.Frame(dir_dialog, padding=(15, 0, 15, 12))
        btn_frame.pack(fill="x")
        btn_frame.pack(side="right")
        ttk.Button(btn_frame, text="确定", command=save).pack(side="right", padx=(0, 6))
        ttk.Button(btn_frame, text="取消", command=dir_dialog.destroy).pack(side="right")

        self.root.wait_window(dir_dialog)

    def _show_about(self):
        messagebox.showinfo(
            "关于",
            "EDM Email Processor\n"
            "自动处理 EDM 邮件：提取 SN、生成 HTML 模板、\n"
            "转换 CSV、导入 Unimarketing 联系人列表。",
            parent=self.root,
        )

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")

        # Font definitions — YaHei UI for UI text, Consolas for monospace
        _ui_font = ("Microsoft YaHei UI", 10)
        _ui_font_sm = ("Microsoft YaHei UI", 9)
        _code_font = ("Consolas", 9)
        _code_font_sm = ("Consolas", 8)

        style.configure("TFrame", background="#ffffff")
        style.configure("TLabel", font=_ui_font_sm, background="#ffffff", foreground="#333333")
        style.configure("Title.TLabel", font=("Microsoft YaHei UI", 15, "bold"), foreground="#1a1a1a")
        style.configure("TButton", padding=(10, 5), font=_ui_font_sm)
        style.configure("Action.TButton", font=("Microsoft YaHei UI", 11, "bold"), padding=(16, 8))
        style.configure("Status.TLabel", font=_code_font)
        style.configure("TLabelframe", font=_ui_font)
        style.configure("TLabelframe.Label", font=("Microsoft YaHei UI", 10, "bold"))
        style.configure("TNotebook", font=_ui_font)
        style.configure("TNotebook.Tab", font=_ui_font_sm, padding=(8, 4))
        style.configure("TEntry", font=_ui_font_sm, padding=4)
        style.map("TButton", background=[("active", "#e0e0e0")])
        style.map("Action.TButton", background=[("active", "#3b82f6")])

        # Main container
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)

        # Title
        ttk.Label(main, text="EDM Email Processor", style="Title.TLabel").pack(
            anchor="w", pady=(0, 10)
        )

        # Notebook — two tabs
        self.notebook = ttk.Notebook(main)
        self.notebook.pack(fill="both", expand=True)

        # Tab 1: EDM Processor
        tab1 = ttk.Frame(self.notebook, padding=4)
        self.notebook.add(tab1, text="  EDM Processor  ")
        self._build_processor_tab(tab1)

        # Tab 2: Verify
        tab2 = ttk.Frame(self.notebook, padding=4)
        self.notebook.add(tab2, text="  Verify  ")
        self._build_verify_tab(tab2)

    # ------------------------------------------------------------------
    # Tab 1 — EDM Processor
    # ------------------------------------------------------------------
    def _build_processor_tab(self, container):
        main = ttk.Frame(container)
        main.pack(fill="both", expand=True)

        # Input files
        input_frame = ttk.LabelFrame(main, text="Input Files", padding=8)
        input_frame.pack(fill="x", pady=(0, 6))
        input_frame.columnconfigure(1, weight=1)

        ttk.Label(input_frame, text="MSG File:").grid(row=0, column=0, sticky="w", padx=(0, 8))
        msg_entry = ttk.Entry(input_frame, textvariable=self.msg_var, width=30, state="readonly")
        msg_entry.grid(row=0, column=1, padx=(0, 6), sticky="ew")
        ttk.Button(input_frame, text="Browse...", command=self._browse_msg).grid(row=0, column=2, sticky="e")

        ttk.Label(input_frame, text="XLSX File:").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(4, 0))
        xlsx_entry = ttk.Entry(input_frame, textvariable=self.xlsx_var, width=30, state="readonly")
        xlsx_entry.grid(row=1, column=1, padx=(0, 6), pady=(4, 0), sticky="ew")
        ttk.Button(input_frame, text="Browse...", command=self._browse_xlsx).grid(row=1, column=2, sticky="e", pady=(4, 0), padx=(0, 3))
        ttk.Button(input_frame, text="Discover", command=self._discover_xlsx).grid(row=1, column=3, sticky="e", pady=(4, 0))

        # Output folder
        output_frame = ttk.LabelFrame(main, text="Output", padding=8)
        output_frame.pack(fill="x", pady=(0, 6))
        output_frame.columnconfigure(1, weight=1)

        ttk.Label(output_frame, text="Folder:").grid(row=0, column=0, sticky="w", padx=(0, 8))
        out_entry = ttk.Entry(output_frame, textvariable=self.output_var, width=30, state="readonly")
        out_entry.grid(row=0, column=1, padx=(0, 6), sticky="ew")
        ttk.Button(output_frame, text="Browse...", command=self._browse_output).grid(row=0, column=2, sticky="e")

        # Buttons — 2 rows
        btn_outer = ttk.Frame(main)
        btn_outer.pack(fill="x", pady=(0, 10))

        btn_row1 = ttk.Frame(btn_outer)
        btn_row1.pack(fill="x", pady=(0, 2))
        btn_row1.columnconfigure(0, weight=1)
        btn_row1.columnconfigure(1, weight=2)
        btn_row1.columnconfigure(2, weight=1)

        self.process_btn = ttk.Button(
            btn_row1, text="  Process  ", command=self._run_process, style="Action.TButton"
        )
        self.process_btn.grid(row=0, column=1, sticky="ew")

        self.open_btn = ttk.Button(
            btn_row1, text="Open Output Folder", command=self._open_folder, state="disabled"
        )
        self.open_btn.grid(row=0, column=2, sticky="e")

        btn_row2 = ttk.Frame(btn_outer)
        btn_row2.pack(fill="x", pady=(2, 0))
        btn_row2.columnconfigure(0, weight=1)
        btn_row2.columnconfigure(1, weight=1)

        self.import_test_btn = ttk.Button(
            btn_row2, text="Import Test List", command=self._import_test_list, state="disabled"
        )
        self.import_test_btn.grid(row=0, column=0, padx=(0, 6), sticky="ew")

        self.import_formal_btn = ttk.Button(
            btn_row2, text="Import Formal List", command=self._import_formal_list, state="disabled"
        )
        self.import_formal_btn.grid(row=0, column=1, padx=(6, 0), sticky="ew")

        # Log
        log_frame = ttk.LabelFrame(main, text="Log", padding=8)
        log_frame.pack(fill="both", expand=True)

        self.log_text = tk.Text(log_frame, height=12, wrap="word", state="disabled",
                                font=("Consolas", 10), bg="#f5f5f5", fg="#222222",
                                insertborderwidth=0, relief="flat")
        self.log_text.pack(fill="both", expand=True)

    # ------------------------------------------------------------------
    # Tab 2 — Verify
    # ------------------------------------------------------------------
    def _build_verify_tab(self, container):
        main = ttk.Frame(container)
        main.pack(fill="both", expand=True)

        # Search inputs — SN and/or Filename (either works)
        sn_frame = ttk.LabelFrame(main, text="Search (SN or Filename)", padding=8)
        sn_frame.pack(fill="x", pady=(0, 10))
        sn_frame.columnconfigure(1, weight=1)

        self.verify_sn_var = tk.StringVar()
        ttk.Label(sn_frame, text="SN:").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=(0, 4))
        sn_entry = ttk.Entry(sn_frame, textvariable=self.verify_sn_var, width=30)
        sn_entry.grid(row=0, column=1, padx=(0, 6), sticky="ew", pady=(0, 4))

        self.verify_filename_var = tk.StringVar()
        ttk.Label(sn_frame, text="Filename:").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(0, 4))
        fn_entry = ttk.Entry(sn_frame, textvariable=self.verify_filename_var, width=30)
        fn_entry.grid(row=1, column=1, padx=(0, 6), sticky="ew", pady=(0, 4))

        ttk.Button(sn_frame, text="Discover", command=self._verify_discover)\
            .grid(row=0, column=2, rowspan=2, sticky="ns", padx=(0, 0))

        # XLSX path (discovered or manual browse)
        xlsx_frame = ttk.LabelFrame(main, text="XLSX File", padding=8)
        xlsx_frame.pack(fill="x", pady=(0, 10))
        xlsx_frame.columnconfigure(1, weight=1)

        self.verify_xlsx_var = tk.StringVar()
        ttk.Label(xlsx_frame, text="Path:").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Entry(xlsx_frame, textvariable=self.verify_xlsx_var, width=50, state="readonly").grid(
            row=0, column=1, padx=(0, 6), sticky="ew"
        )
        ttk.Button(xlsx_frame, text="Browse", command=self._browse_verify_xlsx).grid(row=0, column=2, sticky="e")

        # List info (formal lists found)
        list_frame = ttk.LabelFrame(main, text="Formal Lists (Select One)", padding=8)
        list_frame.pack(fill="x", pady=(0, 10))
        list_frame.columnconfigure(0, weight=1)

        self._verify_lists = []  # list of (list_id, title)
        self._verify_selected_list_id = None

        self.verify_list_box = tk.Listbox(
            list_frame, height=5, font=("Consolas", 10), selectmode="single"
        )
        self.verify_list_box.pack(fill="x")
        self.verify_list_box.bind("<<ListboxSelect>>", self._on_list_select)

        # Verify button
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill="x", pady=(0, 10))
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)

        self.verify_btn = ttk.Button(
            btn_frame, text="  Verify (Email Only)  ", command=self._verify_run, style="Action.TButton"
        )
        self.verify_btn.grid(row=0, column=0, padx=(0, 6), sticky="ew")

        self.deep_verify_btn = ttk.Button(
            btn_frame, text="  Deep Verify (All Fields)  ", command=self._deep_verify_run, style="Action.TButton"
        )
        self.deep_verify_btn.grid(row=0, column=1, padx=(6, 0), sticky="ew")

        # Verify log
        vlog_frame = ttk.LabelFrame(main, text="Verification Log", padding=8)
        vlog_frame.pack(fill="both", expand=True)

        self.verify_log_text = tk.Text(vlog_frame, height=16, wrap="word", state="disabled",
                                        font=("Consolas", 10), bg="#f5f5f5", fg="#222222",
                                        insertborderwidth=0, relief="flat")
        self.verify_log_text.pack(fill="both", expand=True)

    def _browse_msg(self):
        path = filedialog.askopenfilename(
            title="Select MSG File",
            filetypes=[("MSG files", "*.msg"), ("All files", "*.*")],
        )
        if path:
            self.msg_var.set(path)

    def _browse_xlsx(self):
        path = filedialog.askopenfilename(
            title="Select XLSX File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.xlsx_var.set(path)
            self._enable_import_buttons()

    def _discover_xlsx(self):
        """Extract SN from MSG, then search the configured directory for an XLSX file."""
        msg_path = self.msg_var.get().strip()
        if not msg_path or not os.path.isfile(msg_path):
            messagebox.showinfo("Discover", "请先选择 MSG 文件，工具将根据邮件主题中的 SN 号自动检索 XLSX 文件。")
            return

        # Extract SN from MSG
        try:
            from extract_msg import Message as MsgParser
            msg = MsgParser(msg_path)
            subject = msg.subject or ""
            sn = extract_sn(subject)
            if not sn:
                sn = extract_sn(msg_path)
            msg.close()
        except Exception:
            sn = extract_sn(msg_path)

        if not sn:
            messagebox.showinfo("Discover", "无法从 MSG 文件中提取 SN 号，请手动选择 XLSX 文件。")
            return

        # Extract xlsx filename hint from MSG body (SharePoint URL)
        filename_hint = extract_xlsx_filename_from_msg(msg_path)

        search_dir = _load_xlsx_search_dir()
        result = discover_xlsx(sn, search_dir, filename_hint=filename_hint)

        if result:
            self.xlsx_var.set(result)
            self._enable_import_buttons()
            hint_info = f"\n文件名匹配：{os.path.basename(result)}" if filename_hint else ""
            messagebox.showinfo("Discover", f"已找到 XLSX 文件：\n{result}{hint_info}")
        else:
            messagebox.showinfo(
                "Discover",
                f"未找到匹配 {sn} 的 XLSX 文件。\n\n检索目录：\n{search_dir}\n\n请手动选择 XLSX 文件。"
            )

    def _browse_output(self):
        path = filedialog.askdirectory(title="Select Output Folder")
        if path:
            self.output_var.set(path)

    def _run_process(self):
        msg_path = self.msg_var.get().strip()
        xlsx_path = self.xlsx_var.get().strip()
        output_base = self.output_var.get().strip() or DEFAULT_OUTPUT_BASE

        if not msg_path or not os.path.isfile(msg_path):
            messagebox.showerror("Error", "Please select a valid .msg file.")
            return
        if not xlsx_path or not os.path.isfile(xlsx_path):
            messagebox.showerror("Error", "Please select a valid .xlsx file.")
            return

        self.process_btn.config(state="disabled")
        self.open_btn.config(state="disabled")
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

        # Save process log to central Log directory
        log_d = _ensure_log_dir()
        now = datetime.now().strftime("%Y%m%d%H%M%S")
        process_log = os.path.join(log_d, f"process_{now}.log")
        logger = ProcessLogger(self.root, self.log_text, log_file=process_log)
        config = _load_config()

        def on_done(sn_folder):
            self.last_sn_folder = sn_folder
            self.root.after(0, lambda: self.process_btn.config(state="normal"))
            self.root.after(0, self._enable_menu_open_folder)
            self.root.after(0, lambda: self._show_done_dialog(sn_folder))

        def on_error(msg):
            logger.log(f"ERROR: {msg}")
            self.root.after(0, lambda: self.process_btn.config(state="normal"))
            self.root.after(0, lambda: messagebox.showerror("Error", msg))

        thread = threading.Thread(
            target=_process,
            args=(logger, msg_path, xlsx_path, output_base, config, on_done, on_error),
            daemon=True,
        )
        thread.start()

    def _open_folder(self):
        if self.last_sn_folder:
            os.startfile(self.last_sn_folder)

    def _enable_menu_open_folder(self):
        """Enable the menu 'open output folder' command and the main open button."""
        self._file_menu.entryconfig(0, state="normal")
        self.open_btn.config(state="normal")

    def _show_done_dialog(self, sn_folder):
        """Custom dialog after process completes with 'Open Folder' button."""
        sn_name = os.path.basename(sn_folder)
        win = tk.Toplevel(self.root)
        win.title("完成")
        win.geometry("380x150")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()
        win.focus_set()

        ttk.Label(win, text=f"处理完成：{sn_name}",
                  font=("Microsoft YaHei UI", 11, "bold")).pack(
            pady=(18, 8), anchor="w", padx=20
        )

        ttk.Label(win, text=f"输出目录：{sn_folder}",
                  font=("Consolas", 9), foreground="#555").pack(
            anchor="w", padx=20, pady=(0, 16)
        )

        btn_frame = ttk.Frame(win, padding=(20, 0, 20, 18))
        btn_frame.pack(fill="x")

        ttk.Button(
            btn_frame, text="打开文件夹",
            command=lambda: (os.startfile(sn_folder), win.destroy())
        ).pack(side="right", padx=(8, 0))

        ttk.Button(
            btn_frame, text="关闭",
            command=win.destroy
        ).pack(side="right")

    # ---------------------------------------------------------------------------
    # Unimarketing List Import
    # ---------------------------------------------------------------------------

    def _find_sn(self, path: str) -> str | None:
        """Find SN number in path — tries SN-12345 and SN 12345."""
        sn = extract_sn(path)
        if sn:
            return sn
        match = re.search(r"SN\s+(\d+)", path)
        if match:
            return f"SN-{match.group(1)}"
        return None

    def _enable_import_buttons(self):
        """Enable import buttons when xlsx is available."""
        self.import_test_btn.config(state="normal")
        self.import_formal_btn.config(state="normal")

    def _disable_import_buttons(self):
        """Disable import buttons during processing."""
        self.import_test_btn.config(state="disabled")
        self.import_formal_btn.config(state="disabled")

    def _run_import(self, import_fn, is_formal: bool = False):
        """Run an import function directly, passing logger for logging."""
        xlsx_path = self.xlsx_var.get().strip()
        if not xlsx_path or not os.path.isfile(xlsx_path):
            messagebox.showerror("Error", "Please select an xlsx file first.")
            return

        if is_formal:
            if not messagebox.askokcancel(
                "Confirm Formal Import",
                "警告：请确保当前formal发送队列idle后进行。",
            ):
                return

        self._disable_import_buttons()
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

        logger = ProcessLogger(self.root, self.log_text)
        self._last_import_result = None

        def _do_import():
            logger.log("=== Import Started ===")
            try:
                import_result = import_fn(xlsx_path, logger)
                if import_result:
                    list_id, list_title = import_result
                    self.root.after(0, lambda: setattr(self, '_last_import_result', (list_id, list_title)))

                    # Only auto-verify formal imports (deep verify)
                    if is_formal and _HAS_DEEP_VERIFY and list_id:
                        self._run_deep_verify_after_import(list_id, list_title, xlsx_path, logger)
            except SystemExit as e:
                logger.log(f"ERROR: Import exited with code {e.code}")
            except Exception as e:
                logger.log(f"ERROR: {e}")
            logger.log("")
            logger.log("=== Import Done ===")
            self.root.after(0, self._enable_import_buttons)

        thread = threading.Thread(target=_do_import, daemon=True)
        thread.start()

    def _run_deep_verify_after_import(self, list_id: str, list_title: str, xlsx_path: str, logger):
        """Run deep field-by-field verify after formal import completes."""
        save_dir = _ensure_log_dir()
        now = datetime.now().strftime("%Y%m%d%H%M%S")

        safe_title = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in list_title)[:80]
        log_name = f"log_deepverify_formal_{safe_title}_{now}.log"
        log_path = os.path.join(save_dir, log_name)

        gui_root = self.root
        gui_log_fn = logger.log
        gui_save_dir = save_dir
        gui_log_path = log_path

        class _VL:
            @staticmethod
            def log(msg):
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                line = f"[{ts}] {msg}"
                gui_root.after(0, gui_log_fn, line)
                try:
                    os.makedirs(gui_save_dir, exist_ok=True)
                    with open(gui_log_path, "a", encoding="utf-8") as f:
                        f.write(line + "\n")
                except Exception:
                    pass

        vlogger = _VL()

        try:
            passed, msg = _deep_verify(list_id, xlsx_path, save_dir, vlogger)
            if passed:
                self.root.after(0, lambda: messagebox.showinfo("Formal Import Deep Verified", msg))
            else:
                self.root.after(0, lambda: messagebox.showwarning("Formal Import Deep Verify Failed", msg))
        except Exception as e:
            logger.log(f"[VERIFY] Deep verify error: {e}")
            self.root.after(0, lambda: messagebox.showerror("Deep Verify Error", f"Error: {e}"))

    def _run_verify_after_import(self, list_id: str, list_title: str, xlsx_path: str, logger, is_formal: bool):
        """Verify list contacts after import completes."""
        save_dir = _ensure_log_dir()
        log_type = "formal" if is_formal else "test"
        now = datetime.now().strftime("%Y%m%d%H%M%S")

        # Sanitize list title for filename
        safe_title = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in list_title)[:80]
        log_name = f"log_import{log_type}_{safe_title}_{now}.log"
        log_path = os.path.join(save_dir, log_name)

        def file_logger(msg: str):
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            line = f"[{timestamp}] {msg}"
            logger.log(line)
            try:
                os.makedirs(save_dir, exist_ok=True)
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(line + "\n")
            except Exception:
                pass

        try:
            passed, msg = _verify_list_import(
                list_id, xlsx_path, save_dir, file_logger, is_formal
            )
            if passed:
                self.root.after(0, lambda: messagebox.showinfo(
                    "Import Verified", msg
                ))
            else:
                self.root.after(0, lambda: messagebox.showwarning(
                    "Import Verification Failed", msg
                ))
        except Exception as e:
            logger.log(f"[VERIFY] Error: {e}")
            self.root.after(0, lambda: messagebox.showerror(
                "Verify Error", f"Verification error: {e}"
            ))

    def _do_api_import(self, csv_path: str, list_title: str, attrs, logger):
        """Run the Unimarketing API import pipeline with logger logging.
        Returns (list_id, list_title) on success, None on failure.
        """
        logger.log("[API] creating list...")
        list_id = _um_create_list(list_title, attrs)
        if not list_id:
            logger.log("[API] FAILED to create list")
            return None

        logger.log("[API] list created: " + str(list_id))
        logger.log("[API] creating import task...")
        task_title = f"API导入_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        import_id = _um_create_import_task(list_id, task_title)
        if not import_id:
            logger.log("[API] FAILED to create import task")
            return None

        logger.log("[API] import task created: " + str(import_id))
        logger.log("[API] submitting contacts...")
        contact_count = _um_submit_contacts(import_id, csv_path, attrs)
        if not contact_count:
            logger.log("[API] FAILED to submit contacts")
            return None

        logger.log(f"[API] contacts submitted: {contact_count}")
        logger.log("[API] executing import...")
        if not _um_execute_import(import_id):
            logger.log("[API] FAILED to execute import")
            return None

        logger.log("[API] polling import status...")
        result = _um_poll_import_status(import_id, contact_count)

        logger.log("")
        if result.get("status") in ("导入成功", "execute_succeed"):
            logger.log(f"SUCCESS: Import complete — listId={list_id}, importId={import_id}")
            logger.log(f"  Total: {result.get('total')} | Valid: {result.get('validNum')} | "
                        f"Invalid: {result.get('inValidNum')} | Added: {result.get('addToListSuccessNum')} | "
                        f"New: {result.get('addSuccessNum')} | Updated: {result.get('updateSuccessNum')}")
            return list_id, list_title
        else:
            logger.log(f"FAILED: Import ended with status={result.get('status')}")
            return None

    def _import_test_list(self):
        def fn(xlsx_path, logger):
            config = _load_config()
            test_emails = config.get("test_emails", [])
            if not test_emails:
                logger.log("Error: no test_emails in config.json")
                return
            sn = self._find_sn(xlsx_path)
            if not sn:
                logger.log("Error: no SN number found in xlsx path")
                return
            output_base = self.output_var.get().strip() or DEFAULT_OUTPUT_BASE
            output_dir = os.path.join(output_base, sn, "ImportRAW")
            os.makedirs(output_dir, exist_ok=True)
            xlsx_name = os.path.splitext(os.path.basename(xlsx_path))[0]
            now = datetime.now().strftime("%Y%m%d%H%M%S")
            list_title = f"test_{sn}_{xlsx_name}_{now}"

            logger.log(f"=== Unimarketing Test List Import ===")
            logger.log(f"SN: {sn}")
            logger.log(f"List: {list_title}")
            logger.log(f"Test emails: {test_emails}")
            logger.log("")

            logger.log("[CSV] generating test CSV...")
            csv_path = _um_generate_test_csv(xlsx_path, output_dir, test_emails)
            logger.log(f"[CSV] saved: {os.path.basename(csv_path)}")

            with open(csv_path, encoding="gbk", newline="") as f:
                csv_header = next(csv.reader(f))

            attrs = _um_get_attr_mapping(csv_header)
            if not attrs:
                logger.log("[WARN] No Token/SubId columns found")
            logger.log(f"Attributes: {attrs}")
            logger.log("")

            return self._do_api_import(csv_path, list_title, attrs, logger)

        self._run_import(fn)

    def _import_formal_list(self):
        def fn(xlsx_path, logger):
            sn = self._find_sn(xlsx_path)
            if not sn:
                logger.log("Error: no SN number found in xlsx path")
                return
            output_base = self.output_var.get().strip() or DEFAULT_OUTPUT_BASE
            output_dir = os.path.join(output_base, sn, "ImportRAW")
            os.makedirs(output_dir, exist_ok=True)
            xlsx_name = os.path.splitext(os.path.basename(xlsx_path))[0]
            now = datetime.now().strftime("%Y%m%d%H%M%S")
            list_title = f"formal_{sn}_{xlsx_name}_{now}"

            logger.log(f"=== Unimarketing Formal List Import ===")
            logger.log(f"SN: {sn}")
            logger.log(f"List: {list_title}")
            logger.log("")

            logger.log("[CSV] generating formal CSV...")
            csv_path = _um_generate_formal_csv(xlsx_path, output_dir)
            logger.log(f"[CSV] saved: {os.path.basename(csv_path)}")

            with open(csv_path, encoding="gbk", newline="") as f:
                csv_header = next(csv.reader(f))

            attrs = _um_get_attr_mapping(csv_header)
            if not attrs:
                logger.log("[WARN] No Token/SubId columns found")
            logger.log(f"Attributes: {attrs}")
            logger.log("")

            return self._do_api_import(csv_path, list_title, attrs, logger)

        self._run_import(fn, is_formal=True)

    # ------------------------------------------------------------------
    # Verify Tab methods
    # ------------------------------------------------------------------
    def _verify_log(self, message: str):
        """Write to verify tab log."""
        self.verify_log_text.config(state="normal")
        self.verify_log_text.insert("end", message + "\n")
        self.verify_log_text.see("end")
        self.verify_log_text.config(state="disabled")

    def _browse_verify_xlsx(self):
        path = filedialog.askopenfilename(
            title="Select XLSX File for Verification",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.verify_xlsx_var.set(path)
            # Auto-fill SN from path if not already set
            if not self.verify_sn_var.get().strip():
                sn = self._find_sn(path)
                if sn:
                    self.verify_sn_var.set(sn)

    def _verify_discover(self):
        """Discover xlsx by SN and/or filename, then search lists by SN."""
        sn = self.verify_sn_var.get().strip()
        filename_hint = self.verify_filename_var.get().strip()

        if filename_hint and filename_hint != "*.xlsx":
            if not filename_hint.lower().endswith(".xlsx"):
                filename_hint = filename_hint + ".xlsx"
        else:
            filename_hint = None

        # At least one search condition required
        if not sn and not filename_hint:
            messagebox.showinfo("Verify", "请输入 SN 号码和/或 XLSX 文件名")
            return

        # Normalize SN if provided
        if sn:
            digits = "".join(c for c in sn if c.isdigit())
            if not digits:
                messagebox.showinfo("Verify", "SN 格式不正确（如 SN-56287 或 56287）")
                return
            sn_normalized = sn if "-" in sn else f"SN-{sn}"
        else:
            sn_normalized = None
            digits = None

        # Discover xlsx file
        search_dir = _load_xlsx_search_dir()
        result = discover_xlsx(
            sn_normalized,
            search_dir,
            filename_hint=filename_hint,
        )

        if result:
            self.verify_xlsx_var.set(result)
            self._verify_log(f"[DISCOVER] Found xlsx: {result}")
        else:
            parts = []
            if sn_normalized:
                parts.append(f"SN={sn_normalized}")
            if filename_hint:
                parts.append(f"filename={filename_hint}")
            desc = ", ".join(parts)
            self._verify_log(f"[DISCOVER] No xlsx found for {desc}")
            self._verify_log(f"[DISCOVER] Search directory: {search_dir}")
            messagebox.showinfo(
                "Verify",
                f"未找到匹配的 XLSX 文件（{desc}）。\n\n检索目录：\n{search_dir}"
            )
            return

        # Search lists by SN (only if SN was provided)
        if not _HAS_VERIFY:
            messagebox.showerror("Verify", "verify_list_contacts module not loaded")
            return

        if digits:
            self._verify_log(f"[DISCOVER] Searching lists for SN: {digits}")
            lists = _verify_find_lists_by_sn(digits)
        else:
            self._verify_log("[DISCOVER] SN not provided — skipping list search")
            lists = []

        # Populate formal lists in Listbox
        self._verify_lists = [(lid, t) for lid, t, ty in lists if ty == "formal"]
        # Sort by list_id descending (most recent first)
        self._verify_lists.sort(key=lambda x: int(x[0]), reverse=True)

        self.verify_list_box.delete(0, "end")
        for lid, t in self._verify_lists:
            display = f"[{lid}] {t}"
            self.verify_list_box.insert("end", display)
            self._verify_log(f"  formal listId={lid} title={t}")

        self._verify_selected_list_id = None
        self._verify_log(f"[DISCOVER] Found {len(self._verify_lists)} formal list(s)")

        if not self._verify_lists:
            self._verify_log("[DISCOVER] No formal lists found for this SN")

    def _on_list_select(self, event=None):
        """When user selects a list from the Listbox."""
        sel = self.verify_list_box.curselection()
        if sel:
            idx = sel[0]
            self._verify_selected_list_id = self._verify_lists[idx][0]

    def _verify_run(self):
        """Run verification: compare selected list contacts with xlsx emails."""
        if not _HAS_VERIFY:
            messagebox.showerror("Verify", "verify_list_contacts module not loaded")
            return

        xlsx_path = self.verify_xlsx_var.get().strip()
        if not xlsx_path or not os.path.isfile(xlsx_path):
            messagebox.showerror("Verify", "请先输入 SN 号码并点击 Discover 查找 XLSX 文件")
            return

        if not self._verify_selected_list_id:
            messagebox.showerror("Verify", "请在列表中选择一个 Formal List")
            return

        list_id = self._verify_selected_list_id
        list_title = next((t for lid, t in self._verify_lists if lid == list_id), "")

        # Clear log
        self.verify_log_text.config(state="normal")
        self.verify_log_text.delete("1.0", "end")
        self.verify_log_text.config(state="disabled")

        self.verify_btn.config(state="disabled")

        save_dir = _ensure_log_dir()
        now = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_title = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in list_title)[:60]
        log_name = f"log_verify_{list_id}_{now}.log"
        log_path = os.path.join(save_dir, log_name)

        # Logger object with .log() method — use @staticmethod to avoid self param
        # Logger object with .log() method — capture in closure
        gui_root = self.root
        gui_verify_log_fn = self._verify_log
        gui_save_dir = save_dir
        gui_log_path = log_path

        class _VL:
            @staticmethod
            def log(msg):
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                line = f"[{ts}] {msg}"
                gui_root.after(0, gui_verify_log_fn, line)
                try:
                    os.makedirs(gui_save_dir, exist_ok=True)
                    with open(gui_log_path, "a", encoding="utf-8") as f:
                        f.write(line + "\n")
                except Exception:
                    pass

        vlogger = _VL()

        def _do():
            try:
                passed, msg = _verify_list_import(
                    list_id, xlsx_path, save_dir, vlogger, is_formal=True
                )
                if passed:
                    self.root.after(0, lambda: messagebox.showinfo("验证成功", msg))
                else:
                    self.root.after(0, lambda: messagebox.showwarning("验证失败", msg))
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.root.after(0, lambda: messagebox.showerror("验证错误", f"Error: {e}"))
            self.root.after(0, lambda: self.verify_btn.config(state="normal"))

        thread = threading.Thread(target=_do, daemon=True)
        thread.start()

    def _deep_verify_run(self):
        """Run deep field-by-field verification."""
        if not _HAS_DEEP_VERIFY:
            messagebox.showerror("Deep Verify", "deep_verify_list module not loaded")
            return

        xlsx_path = self.verify_xlsx_var.get().strip()
        if not xlsx_path or not os.path.isfile(xlsx_path):
            messagebox.showerror("Deep Verify", "请先输入 SN 号码并点击 Discover 查找 XLSX 文件")
            return

        if not self._verify_selected_list_id:
            messagebox.showerror("Deep Verify", "请在列表中选择一个 Formal List")
            return

        list_id = self._verify_selected_list_id
        list_title = next((t for lid, t in self._verify_lists if lid == list_id), "")

        # Clear log
        self.verify_log_text.config(state="normal")
        self.verify_log_text.delete("1.0", "end")
        self.verify_log_text.config(state="disabled")

        self.verify_btn.config(state="disabled")
        self.deep_verify_btn.config(state="disabled")

        save_dir = _ensure_log_dir()
        now = datetime.now().strftime("%Y%m%d%H%M%S")
        safe_title = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in list_title)[:60]
        log_name = f"log_deepverify_{list_id}_{now}.log"
        log_path = os.path.join(save_dir, log_name)

        gui_root = self.root
        gui_verify_log_fn = self._verify_log
        gui_save_dir = save_dir
        gui_log_path = log_path

        class _VL:
            @staticmethod
            def log(msg):
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                line = f"[{ts}] {msg}"
                gui_root.after(0, gui_verify_log_fn, line)
                try:
                    os.makedirs(gui_save_dir, exist_ok=True)
                    with open(gui_log_path, "a", encoding="utf-8") as f:
                        f.write(line + "\n")
                except Exception:
                    pass

        vlogger = _VL()

        def _do():
            try:
                passed, msg = _deep_verify(
                    list_id, xlsx_path, save_dir, vlogger
                )
                if passed:
                    self.root.after(0, lambda: messagebox.showinfo("深验证成功", msg))
                else:
                    self.root.after(0, lambda: messagebox.showwarning("深验证完成", msg))
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.root.after(0, lambda: messagebox.showerror("深验证错误", f"Error: {e}"))
            self.root.after(0, lambda: self.verify_btn.config(state="normal"))
            self.root.after(0, lambda: self.deep_verify_btn.config(state="normal"))

        thread = threading.Thread(target=_do, daemon=True)
        thread.start()

    def run(self):
        self.root.mainloop()


# ---------------------------------------------------------------------------
# Helpers for config display
# ---------------------------------------------------------------------------

def _format_config_status(label: str, raw: str | None) -> str:
    if raw is None:
        return f"{label}: not found"
    try:
        data = json.loads(raw)
        if label == "config.json":
            emails = data.get("test_emails", [])
            parts = [f"{label}: {len(emails)} email(s)"]
            for e in emails:
                parts.append(f"  → {e}")
            return "\n".join(parts)
        else:
            count = len(data) if isinstance(data, list) else len(data)
            return f"{label}: {count} token(s)"
    except json.JSONDecodeError:
        return f"{label}: invalid JSON"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app = EDMGUI()
    app.run()
